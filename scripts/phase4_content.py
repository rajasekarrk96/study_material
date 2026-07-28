"""
phase4_content.py
Fills stubs for:
  _11_mysql  (1)
  _16_selenium (24)
  _08_java (21)
  _09_c    (16)
  _10_cpp  (13)
"""
import os

BASE = r'd:\My Drive\all files\PROJECT FILES\notes\docs\curriculum'
written = 0

def write(course_dir, fname, content):
    global written
    path = os.path.join(BASE, course_dir, fname)
    os.makedirs(os.path.dirname(path), exist_ok=True)
    with open(path, 'w', encoding='utf-8') as f:
        f.write(content)
    print(f'  [WRITE] {fname}')
    written += 1

def fm(lid, title, course, mod, mod_title, les, diff, tags, dur=60):
    tag_str = ', '.join(f'"{t}"' for t in tags)
    return f'''---
id: "{lid}"
title: "{title}"
course: "{course}"
module: {mod}
module_title: "{mod_title}"
lesson: {les}
version: "2.0"
difficulty: "{diff}"
duration_minutes: {dur}
tags: [{tag_str}]
prerequisites: []
lab_required: true
---

# {title}

'''

# ═══════════════════════════════════════════════════════════════
# MYSQL — 1 remaining
# ═══════════════════════════════════════════════════════════════
print('='*60)
print('MYSQL — 1 lesson')
print('='*60)

write('_11_mysql', '_05_01_database_architecture_and_relational_concepts.md',
fm('05_01','Database Architecture and Relational Concepts','MySQL',1,'MySQL Foundations',1,'beginner',
   ['RDBMS','relational-model','ACID','table','row','column','schema','primary-key','foreign-key','normalization','SQL']) + '''
## What is a Relational Database?

A **Relational Database Management System (RDBMS)** organises data into **tables** (relations) with rows (tuples) and columns (attributes), enforcing relationships through keys.

### Key Concepts

| Term | Definition |
|---|---|
| **Table** | 2-D structure with rows and columns |
| **Row** | One record (instance of an entity) |
| **Column** | One attribute with a defined data type |
| **Primary Key** | Uniquely identifies each row |
| **Foreign Key** | References PK in another table (enforces referential integrity) |
| **Schema** | Blueprint of the database structure |
| **Index** | Data structure that speeds up queries |
| **View** | Virtual table defined by a SELECT query |

## MySQL Architecture

```
Client Layer       → mysql CLI, MySQL Workbench, application drivers
   ↓
Connection/Thread  → Each client gets a thread; connection pool
   ↓
SQL Interface      → Parser → Optimizer → Executor
   ↓
Storage Engine     → InnoDB (default), MyISAM, MEMORY, CSV
   ↓
File System        → Data files (.ibd), redo log, undo log, binary log
```

### Storage Engines Comparison

| Feature | InnoDB | MyISAM |
|---|---|---|
| ACID | Yes | No |
| Transactions | Yes | No |
| Foreign Keys | Yes | No |
| Row-level locking | Yes | Table-level only |
| Full-text search | Yes (5.6+) | Yes |
| Use case | General OLTP | Read-heavy archives |

## ACID Properties

```
A — Atomicity   : All operations in a transaction succeed or ALL are rolled back
C — Consistency : DB moves from one valid state to another
I — Isolation   : Concurrent transactions don't interfere
D — Durability  : Committed data survives crashes (written to disk)
```

## SQL Categories

| Category | Commands | Purpose |
|---|---|---|
| **DDL** — Data Definition | CREATE, ALTER, DROP, TRUNCATE | Structure |
| **DML** — Data Manipulation | INSERT, UPDATE, DELETE | Data |
| **DQL** — Data Query | SELECT | Retrieve |
| **DCL** — Data Control | GRANT, REVOKE | Permissions |
| **TCL** — Transaction Control | BEGIN, COMMIT, ROLLBACK, SAVEPOINT | Transactions |

## Connecting to MySQL

```bash
# CLI
mysql -u root -p
mysql -u root -p mydb

# Inside MySQL shell
SHOW DATABASES;
USE mydb;
SHOW TABLES;
DESCRIBE employees;
SHOW CREATE TABLE employees;
```

## Lab Exercise
1. Install MySQL 8.0, create a database `school`, create tables `students` and `courses`
2. Insert 10 rows into each table and verify with `SELECT * FROM ...`
3. Use `DESCRIBE` and `SHOW CREATE TABLE` to inspect your schema
''')

# ═══════════════════════════════════════════════════════════════
# SELENIUM — 24 lessons
# ═══════════════════════════════════════════════════════════════
print()
print('='*60)
print('SELENIUM — 24 lessons')
print('='*60)
S = '_16_selenium'

write(S,'_16_01_01_selenium_introduction_and_setup.md',
fm('16_01_01','Selenium Introduction and Setup','Selenium',1,'Selenium Fundamentals',1,'beginner',
   ['selenium','webdriver','browser-automation','pip','chromedriver','geckodriver','selenium-manager']) + '''
## What is Selenium?

**Selenium** is an open-source browser automation framework that allows you to programmatically control web browsers — click buttons, fill forms, extract data, and validate UI behaviour.

### Selenium Suite Components

| Tool | Purpose |
|---|---|
| **Selenium WebDriver** | Core API to control browsers |
| **Selenium Grid** | Distributed test execution across machines/browsers |
| **Selenium IDE** | Record & playback browser extension (no-code) |

## Installation

```bash
# Install Selenium
pip install selenium

# Selenium 4.6+ includes Selenium Manager (auto-downloads drivers!)
# No manual chromedriver install needed for Chrome/Firefox/Edge
```

## First Script

```python
from selenium import webdriver
from selenium.webdriver.common.by import By
import time

# Chrome (headless option)
options = webdriver.ChromeOptions()
# options.add_argument("--headless")  # uncomment for headless

driver = webdriver.Chrome(options=options)

try:
    driver.get("https://www.google.com")
    print(driver.title)   # Google

    search_box = driver.find_element(By.NAME, "q")
    search_box.send_keys("Selenium Python")
    search_box.submit()
    time.sleep(2)
    print(driver.title)
finally:
    driver.quit()   # always quit to free resources
```

## Browser Options

```python
# Chrome
from selenium.webdriver.chrome.options import Options as ChromeOptions
opts = ChromeOptions()
opts.add_argument("--start-maximized")
opts.add_argument("--disable-notifications")
opts.add_experimental_option("detach", True)   # keep browser open after script
driver = webdriver.Chrome(options=opts)

# Firefox
from selenium.webdriver.firefox.options import Options as FirefoxOptions
opts = FirefoxOptions()
driver = webdriver.Firefox(options=opts)

# Edge
driver = webdriver.Edge()
```

## Lab Exercise
1. Set up a virtual environment, install selenium, verify `selenium.__version__`
2. Write a script that opens `https://example.com`, prints the title and current URL
3. Navigate to Wikipedia, search for "Python programming", print the first paragraph
''')

write(S,'_16_01_02_webdriver_core_and_browser_control.md',
fm('16_01_02','WebDriver Core and Browser Control','Selenium',1,'Selenium Fundamentals',2,'beginner',
   ['get','back','forward','refresh','title','current-url','window-size','maximize','screenshot','quit','close']) + '''
## WebDriver Navigation

```python
driver.get("https://example.com")    # open URL (blocks until loaded)
driver.back()                         # browser Back button
driver.forward()                      # browser Forward button
driver.refresh()                      # reload page
```

## Browser Properties

```python
driver.title           # page title string
driver.current_url     # current URL string
driver.page_source     # full HTML source
```

## Window Management

```python
driver.maximize_window()
driver.minimize_window()
driver.set_window_size(1920, 1080)
driver.set_window_position(0, 0)
driver.get_window_size()    # {'width': 1920, 'height': 1080}
```

## Screenshots

```python
# Capture full page screenshot
driver.save_screenshot("screenshot.png")

# Get as bytes (for embedding in reports)
png_bytes = driver.get_screenshot_as_png()

# Screenshot of specific element
element = driver.find_element(By.ID, "header")
element.screenshot("header.png")
```

## Cookies

```python
driver.get_cookies()                          # list of all cookies
driver.get_cookie("session_id")              # specific cookie
driver.add_cookie({"name": "token", "value": "abc123"})
driver.delete_cookie("session_id")
driver.delete_all_cookies()
```

## Execute Script

```python
# Scroll to bottom of page
driver.execute_script("window.scrollTo(0, document.body.scrollHeight);")

# Get value not accessible via Selenium
inner_text = driver.execute_script("return arguments[0].innerText;", element)
```

## Quit vs Close

```python
driver.close()   # closes current window/tab only
driver.quit()    # quits entire browser + kills WebDriver process
# Always use quit() at end of test!
```

## Lab Exercise
1. Script that navigates back and forward through 3 pages, verifying the URL each time
2. Save a screenshot before and after clicking a button
3. Set a custom cookie and verify it persists on page reload
''')

write(S,'_16_01_03_locator_strategies.md',
fm('16_01_03','Locator Strategies','Selenium',1,'Selenium Fundamentals',3,'beginner',
   ['By','find_element','find_elements','ID','NAME','CLASS_NAME','TAG_NAME','LINK_TEXT','PARTIAL_LINK_TEXT','CSS_SELECTOR','XPATH']) + '''
## By Locator Types

```python
from selenium.webdriver.common.by import By

# By ID — fastest, most reliable
element = driver.find_element(By.ID, "username")

# By NAME
element = driver.find_element(By.NAME, "q")

# By CLASS_NAME (only first class)
elements = driver.find_elements(By.CLASS_NAME, "product-card")

# By TAG NAME
all_links = driver.find_elements(By.TAG_NAME, "a")

# By LINK TEXT (exact text of <a>)
driver.find_element(By.LINK_TEXT, "Sign In")

# By PARTIAL LINK TEXT
driver.find_element(By.PARTIAL_LINK_TEXT, "Sign")

# By CSS SELECTOR — flexible, fast
driver.find_element(By.CSS_SELECTOR, "#login-form input[type='email']")
driver.find_elements(By.CSS_SELECTOR, ".product-card > .price")

# By XPATH — most powerful, slowest
driver.find_element(By.XPATH, "//input[@id='username']")
driver.find_elements(By.XPATH, "//table[@class='data']//tr")
```

## find_element vs find_elements

```python
# find_element — returns ONE element; NoSuchElementException if not found
element = driver.find_element(By.ID, "header")

# find_elements — returns LIST; empty list if none found (no exception)
elements = driver.find_elements(By.CLASS_NAME, "item")
print(len(elements))    # 0 if not found
```

## Choosing the Right Locator

| Priority | Strategy | Why |
|---|---|---|
| 1 | `ID` | Unique, fast, semantic |
| 2 | `NAME` | Common on form fields |
| 3 | `CSS_SELECTOR` | Fast, readable, powerful |
| 4 | `XPATH` | When CSS can't reach it |
| 5 | `CLASS_NAME` | When unique enough |
| ✗ | `TAG_NAME` | Too broad |
| ✗ | `LINK_TEXT` | Breaks on text changes |

## Relative Locators (Selenium 4)

```python
from selenium.webdriver.support.relative_locator import locate_with

email = driver.find_element(By.ID, "email")

# Find password field BELOW email field
password = driver.find_element(
    locate_with(By.TAG_NAME, "input").below(email)
)

# More relative locators
.above(element)
.to_left_of(element)
.to_right_of(element)
.near(element)
```

## Lab Exercise
1. Locate 5 elements on a login page using 5 different By strategies
2. Use `find_elements` to count all links on a Wikipedia page
3. Demonstrate relative locators by finding a label next to an input
''')

write(S,'_16_01_04_xpath_and_css_selectors.md',
fm('16_01_04','XPath and CSS Selectors','Selenium',1,'Selenium Fundamentals',4,'intermediate',
   ['xpath','css-selector','axes','predicates','contains','text()','attribute','parent','sibling','nth-child']) + '''
## CSS Selector Syntax

```css
/* By element type */
input
button

/* By ID */
#username
input#username

/* By class */
.btn-primary
button.btn.btn-primary    /* multiple classes */

/* By attribute */
input[type="email"]
input[placeholder="Enter email"]
[data-testid="submit-btn"]

/* Child selectors */
form > input           /* direct child */
.container input       /* any descendant */

/* Nth child */
li:nth-child(2)        /* second li */
li:first-child
li:last-child
li:nth-of-type(3)

/* Chaining */
#login-form input[type="password"]
.product-list > .product-card > .price
```

## XPath Syntax

```xpath
/* Basic */
//input                          all input elements
//input[@id="username"]          input with id=username
//input[@type="email"]           input with type=email

/* Text content */
//button[text()="Submit"]
//a[contains(text(), "Login")]
//h1[normalize-space()="Home"]

/* Contains */
//input[contains(@class, "form-control")]
//div[contains(@id, "product-")]

/* Starts-with */
//input[starts-with(@name, "user")]

/* Axes */
//label[@for="email"]/following-sibling::input
//td[text()="Price"]/following-sibling::td
//input[@id="email"]/parent::div
//tr/td[1]                      first column of each row

/* Index (1-based!) */
(//li[@class="item"])[1]
(//tr)[last()]

/* AND / OR */
//input[@type="text" and @required]
//button[@type="submit" or @type="button"]
```

## Practical Examples

```python
# Find table cell containing "Active" in same row as specific name
driver.find_element(By.XPATH,
    "//tr[td[text()='John Doe']]/td[contains(@class,'status')]")

# Find button that comes after a specific heading
driver.find_element(By.XPATH,
    "//h2[text()='Payment Details']/following::button[@type='submit']")

# Dynamic ID that starts with known prefix
driver.find_element(By.CSS_SELECTOR, "[id^='react-select-']")
```

## CSS vs XPath — When to Use

| Feature | CSS | XPath |
|---|---|---|
| Speed | Faster | Slightly slower |
| Readability | More readable | Complex |
| Parent traversal | Not supported | Supported |
| Text matching | Not supported | Supported |
| Browser support | Universal | Universal |

## Lab Exercise
1. Extract all product names from a table using XPath axes
2. Write CSS selector for: input with class containing "form" inside a div with id "checkout"
3. Use XPath to find a button that follows a specific label text
''')

write(S,'_16_01_05_web_element_interactions.md',
fm('16_01_05','Web Element Interactions','Selenium',1,'Selenium Fundamentals',5,'beginner',
   ['click','send_keys','clear','text','get_attribute','is_displayed','is_enabled','is_selected','submit','value']) + '''
## Core Element Methods

```python
element = driver.find_element(By.ID, "username")

# Clicking
element.click()

# Typing
element.send_keys("myusername")
element.clear()                         # clear existing text
element.send_keys("new text")

# Form submit
form = driver.find_element(By.ID, "login-form")
form.submit()

# Read content
print(element.text)                     # visible text
print(element.get_attribute("value"))   # input value
print(element.get_attribute("href"))    # link href
print(element.get_attribute("class"))   # class attribute
print(element.get_attribute("innerHTML"))
print(element.get_attribute("outerHTML"))

# State checks
element.is_displayed()   # True if visible
element.is_enabled()     # True if not disabled
element.is_selected()    # True if checkbox/radio checked
```

## Special Keys

```python
from selenium.webdriver.common.keys import Keys

element.send_keys(Keys.RETURN)       # Enter
element.send_keys(Keys.TAB)          # Tab
element.send_keys(Keys.ESCAPE)       # Esc
element.send_keys(Keys.BACKSPACE)    # Backspace
element.send_keys(Keys.CONTROL, "a") # Ctrl+A (select all)
element.send_keys(Keys.CONTROL, "c") # Ctrl+C
element.send_keys(Keys.HOME)
element.send_keys(Keys.END)
element.send_keys(Keys.PAGE_DOWN)
element.send_keys(Keys.ARROW_DOWN)
```

## CSS Properties and Dimensions

```python
# CSS value
color = element.value_of_css_property("color")
font_size = element.value_of_css_property("font-size")

# Location and size
location = element.location      # {'x': 100, 'y': 200}
size = element.size              # {'width': 300, 'height': 50}
rect = element.rect              # {'x', 'y', 'width', 'height'}
```

## Checkbox and Radio Buttons

```python
checkbox = driver.find_element(By.ID, "agree-terms")

if not checkbox.is_selected():
    checkbox.click()   # check it

# Verify state
assert checkbox.is_selected()

# Radio buttons
radio = driver.find_element(By.XPATH, "//input[@type='radio' and @value='monthly']")
radio.click()
```

## Lab Exercise
1. Automate a login form: clear fields, type credentials, click submit, verify redirect
2. Interact with a form that has checkboxes, radio buttons, and a text area
3. Verify a button is disabled before form validation and enabled after
''')

write(S,'_16_02_01_implicit_and_explicit_waits.md',
fm('16_02_01','Implicit and Explicit Waits','Selenium',2,'Waits and Synchronisation',1,'intermediate',
   ['implicitly_wait','WebDriverWait','expected_conditions','EC','visibility','presence','clickable','timeout','polling']) + '''
## Why Waits Are Necessary

Modern web apps load content dynamically (AJAX, React, Vue). Without waits, Selenium may try to interact with elements that haven't appeared yet, causing `NoSuchElementException` or `ElementNotInteractableException`.

## Implicit Wait

```python
# Set once — applies to ALL find_element calls for the session
driver.implicitly_wait(10)   # wait up to 10 seconds for element to appear

# Selenium polls the DOM at ~500ms intervals
# Use sparingly — slows down tests when elements genuinely don't exist
```

## Explicit Wait (Recommended)

```python
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC

wait = WebDriverWait(driver, timeout=10)

# Wait for element to be visible
element = wait.until(
    EC.visibility_of_element_located((By.ID, "success-message"))
)

# Wait for element to be clickable (visible AND enabled)
btn = wait.until(
    EC.element_to_be_clickable((By.CSS_SELECTOR, "#submit-btn"))
)
btn.click()

# Wait for element to disappear
wait.until(
    EC.invisibility_of_element_located((By.CLASS_NAME, "loading-spinner"))
)
```

## Common Expected Conditions

```python
from selenium.webdriver.support import expected_conditions as EC

EC.title_contains("Dashboard")
EC.title_is("Home Page")
EC.url_contains("/dashboard")
EC.url_to_be("https://example.com/dashboard")

EC.presence_of_element_located(locator)   # in DOM, may not be visible
EC.visibility_of_element_located(locator)  # visible (display != none)
EC.element_to_be_clickable(locator)        # visible AND enabled
EC.invisibility_of_element_located(locator) # hidden/removed

EC.presence_of_all_elements_located(locator)  # returns list
EC.text_to_be_present_in_element(locator, "text")
EC.element_to_be_selected(element)
EC.staleness_of(element)    # element removed from DOM (after navigation)
EC.number_of_windows_to_be(2)
EC.alert_is_present()
```

## Custom Wait Condition

```python
def element_has_css_class(locator, css_class):
    def condition(driver):
        element = driver.find_element(*locator)
        classes = element.get_attribute("class")
        return css_class in classes.split()
    return condition

wait.until(element_has_css_class((By.ID, "status"), "active"))
```

## Implicit vs Explicit

| | Implicit | Explicit |
|---|---|---|
| Scope | All find_element calls | Specific condition |
| Flexibility | Low | High |
| Timeout | Single global | Per-wait |
| Conditions | Presence only | Any condition |
| Recommendation | Avoid mixing with explicit | Preferred |

> ⚠️ Never mix implicit and explicit waits — leads to unpredictable timeouts.

## Lab Exercise
1. Use `element_to_be_clickable` to wait for a button that appears after 3 seconds
2. Wait for a loading spinner to disappear before asserting page content
3. Write a custom condition that checks element text matches a regex
''')

write(S,'_16_02_02_fluent_waits_and_custom_conditions.md',
fm('16_02_02','Fluent Waits and Custom Conditions','Selenium',2,'Waits and Synchronisation',2,'advanced',
   ['FluentWait','polling','ignored_exceptions','custom-condition','lambda','timeout','NoSuchElementException']) + '''
## FluentWait

`FluentWait` gives fine-grained control over polling frequency and which exceptions to ignore.

```python
from selenium.webdriver.support.wait import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from selenium.common.exceptions import NoSuchElementException, StaleElementReferenceException
from datetime import timedelta

# FluentWait via WebDriverWait parameters
wait = WebDriverWait(
    driver,
    timeout=30,
    poll_frequency=0.5,                  # check every 500ms (default: 500ms)
    ignored_exceptions=[
        NoSuchElementException,
        StaleElementReferenceException,  # ignore stale elements during polling
    ]
)

element = wait.until(
    EC.visibility_of_element_located((By.ID, "dynamic-content"))
)
```

## Custom Wait Conditions with Lambda

```python
# Wait for element count to be at least N
wait.until(lambda d: len(d.find_elements(By.CLASS_NAME, "result-item")) >= 5)

# Wait for element's text to contain expected value
wait.until(lambda d:
    "success" in d.find_element(By.ID, "status").text.lower()
)

# Wait for a specific attribute value
wait.until(lambda d:
    d.find_element(By.ID, "progress").get_attribute("aria-valuenow") == "100"
)

# Wait for URL to change
original_url = driver.current_url
driver.find_element(By.ID, "navigate-btn").click()
wait.until(lambda d: d.current_url != original_url)
```

## Retry Decorator for Flaky Interactions

```python
import time
from functools import wraps
from selenium.common.exceptions import StaleElementReferenceException

def retry_on_stale(retries=3, delay=0.5):
    def decorator(func):
        @wraps(func)
        def wrapper(*args, **kwargs):
            for attempt in range(retries):
                try:
                    return func(*args, **kwargs)
                except StaleElementReferenceException:
                    if attempt == retries - 1:
                        raise
                    time.sleep(delay)
        return wrapper
    return decorator

@retry_on_stale(retries=3)
def click_element(driver, locator):
    driver.find_element(*locator).click()
```

## Wait Until Page is Ready

```python
def wait_for_page_load(driver, timeout=30):
    WebDriverWait(driver, timeout).until(
        lambda d: d.execute_script("return document.readyState") == "complete"
    )

def wait_for_angular(driver, timeout=30):
    WebDriverWait(driver, timeout).until(
        lambda d: d.execute_script(
            "return window.getAllAngularTestabilities && "
            "window.getAllAngularTestabilities().every(t => t.isStable())"
        )
    )
```

## Lab Exercise
1. Implement a `wait_for_ajax(driver)` function that polls until no jQuery AJAX requests are pending
2. Use FluentWait to ignore `StaleElementReferenceException` while waiting for a result table
3. Build a `wait_for_text_change(element, original_text)` custom condition
''')

write(S,'_16_02_03_page_load_strategies.md',
fm('16_02_03','Page Load Strategies','Selenium',2,'Waits and Synchronisation',3,'intermediate',
   ['pageLoad','normal','eager','none','readyState','AJAX','SPA','timeouts','set_page_load_timeout']) + '''
## Page Load Strategies

```python
from selenium.webdriver.chrome.options import Options

options = Options()

# NORMAL (default) — waits for full page load (document.readyState == "complete")
options.page_load_strategy = "normal"

# EAGER — waits for interactive (DOM ready, resources may still load)
options.page_load_strategy = "eager"

# NONE — returns immediately after initial request
options.page_load_strategy = "none"

driver = webdriver.Chrome(options=options)
```

## Timeouts Configuration

```python
driver.set_page_load_timeout(30)     # max seconds to wait for page to load
driver.set_script_timeout(10)        # max seconds for execute_async_script
driver.implicitly_wait(5)            # implicit wait for element presence

# Via options (Selenium 4 preferred)
from selenium.webdriver.chrome.options import Options
options = Options()
options.set_capability("timeouts", {
    "pageLoad": 30000,   # milliseconds
    "script": 10000,
    "implicit": 0,
})
```

## Handling Slow AJAX Pages

```python
def wait_for_ajax(driver, timeout=30):
    """Wait until jQuery AJAX requests complete"""
    from selenium.webdriver.support.wait import WebDriverWait
    WebDriverWait(driver, timeout).until(
        lambda d: d.execute_script("return jQuery.active == 0")
    )

def wait_for_network_idle(driver, timeout=30):
    """Wait for React/Vue single-page apps to finish rendering"""
    WebDriverWait(driver, timeout).until(
        lambda d: d.execute_script(
            "return document.readyState === 'complete'"
        )
    )
```

## Lab Exercise
1. Compare page load times for `normal`, `eager`, `none` strategies on a heavy website
2. Write a `navigate_and_wait(driver, url)` that uses `none` strategy then polls for readiness
3. Handle a page that loads content via AJAX 2 seconds after initial DOM load
''')

write(S,'_16_03_01_action_chains.md',
fm('16_03_01','Action Chains','Selenium',3,'Advanced Interactions',1,'intermediate',
   ['ActionChains','click_and_hold','drag_and_drop','move_to_element','hover','context_click','double_click','key_down','perform']) + '''
## ActionChains Overview

`ActionChains` build a sequence of low-level browser actions (mouse movements, clicks, key presses) that are performed in order.

```python
from selenium.webdriver.common.action_chains import ActionChains

action = ActionChains(driver)
```

## Mouse Actions

```python
element = driver.find_element(By.ID, "target")

# Hover (mouse over)
action.move_to_element(element).perform()

# Click types
action.click(element).perform()
action.double_click(element).perform()
action.context_click(element).perform()   # right-click

# Click and hold (drag start)
action.click_and_hold(element).perform()
action.release().perform()

# Move by offset from element
action.move_to_element_with_offset(element, 10, 20).perform()

# Move by offset from current position
action.move_by_offset(100, 0).perform()
```

## Drag and Drop

```python
source = driver.find_element(By.ID, "drag-item")
target = driver.find_element(By.ID, "drop-zone")

# Method 1: drag_and_drop
action.drag_and_drop(source, target).perform()

# Method 2: manual (more reliable for some apps)
action.click_and_hold(source) \
      .move_to_element(target) \
      .release() \
      .perform()
```

## Keyboard Actions

```python
from selenium.webdriver.common.keys import Keys

# Key combinations
action.key_down(Keys.CONTROL) \
      .send_keys("a") \
      .key_up(Keys.CONTROL) \
      .perform()   # Ctrl+A

# Type in focused element
action.send_keys("Hello World").perform()

# Tab through form fields
action.send_keys(Keys.TAB).perform()
```

## Chaining Actions

```python
# Hover over menu, wait for submenu, click submenu item
menu = driver.find_element(By.ID, "nav-products")
submenu_item = driver.find_element(By.ID, "nav-laptops")

ActionChains(driver) \
    .move_to_element(menu) \
    .pause(0.5) \
    .move_to_element(submenu_item) \
    .click() \
    .perform()
```

## Lab Exercise
1. Automate a drag-and-drop kanban board (move card from "To Do" to "In Progress")
2. Open a dropdown navigation menu by hovering, then click a submenu link
3. Select all text in an input field using keyboard shortcut and replace it
''')

write(S,'_16_03_02_dropdown_and_select_handling.md',
fm('16_03_02','Dropdown and Select Handling','Selenium',3,'Advanced Interactions',2,'intermediate',
   ['Select','select_by_value','select_by_visible_text','select_by_index','deselect','options','all_selected_options','multi-select']) + '''
## HTML Select Dropdown

```python
from selenium.webdriver.support.ui import Select

dropdown = driver.find_element(By.ID, "country-select")
select = Select(dropdown)

# Select by visible text
select.select_by_visible_text("India")

# Select by value attribute
select.select_by_value("IN")

# Select by index (0-based)
select.select_by_index(2)

# Get all options
for option in select.options:
    print(option.text, option.get_attribute("value"))

# Get currently selected option
print(select.first_selected_option.text)
print(select.all_selected_options)   # list (for multi-select)
```

## Multi-Select Dropdown

```python
select = Select(driver.find_element(By.ID, "languages"))

select.select_by_visible_text("Python")
select.select_by_visible_text("Java")
select.select_by_visible_text("JavaScript")

# Deselect
select.deselect_by_visible_text("Java")
select.deselect_all()

print([opt.text for opt in select.all_selected_options])
```

## Custom Dropdown (not `<select>`)

Many modern UI frameworks use `<div>` or `<ul>` based dropdowns.

```python
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC

# Click to open
driver.find_element(By.CSS_SELECTOR, ".dropdown-toggle").click()

# Wait for options to appear
wait = WebDriverWait(driver, 10)
wait.until(EC.visibility_of_element_located((By.CSS_SELECTOR, ".dropdown-menu")))

# Click specific option
options = driver.find_elements(By.CSS_SELECTOR, ".dropdown-menu li a")
for option in options:
    if option.text == "Settings":
        option.click()
        break
```

## Lab Exercise
1. Select a date from three separate dropdowns (day, month, year)
2. Verify all options are present in a dropdown and select each one in sequence
3. Handle a Bootstrap dropdown (custom div-based) and select a value
''')

write(S,'_16_03_03_alerts_frames_windows.md',
fm('16_03_03','Alerts Frames and Windows','Selenium',3,'Advanced Interactions',3,'intermediate',
   ['alert','switch_to.alert','accept','dismiss','send_keys','iframe','frame','switch_to.frame','window_handles','switch_to.window']) + '''
## JavaScript Alerts

```python
from selenium.webdriver.support import expected_conditions as EC
from selenium.webdriver.support.ui import WebDriverWait

# Wait for alert to appear
wait = WebDriverWait(driver, 10)
wait.until(EC.alert_is_present())

alert = driver.switch_to.alert

# Alert types
alert.accept()    # OK button
alert.dismiss()   # Cancel button
print(alert.text) # alert message text

# Prompt (alert with input)
alert.send_keys("My answer")
alert.accept()
```

## iFrames

```python
# Switch to frame by index (0-based)
driver.switch_to.frame(0)

# Switch to frame by name or ID attribute
driver.switch_to.frame("iframe-name")

# Switch to frame by WebElement
iframe = driver.find_element(By.CSS_SELECTOR, "iframe.content-frame")
driver.switch_to.frame(iframe)

# Interact with content inside frame
driver.find_element(By.ID, "submit").click()

# Return to main page
driver.switch_to.default_content()

# Switch to parent frame (from nested frame)
driver.switch_to.parent_frame()
```

## Multiple Windows/Tabs

```python
# Open new tab
driver.execute_script("window.open('https://example.com', '_blank');")

# Get all window handles
handles = driver.window_handles
print(handles)   # ['handle1', 'handle2']

main_window = driver.current_window_handle

# Switch to new window
driver.switch_to.window(handles[-1])
print(driver.title)   # Title of new window

# Switch back
driver.switch_to.window(main_window)

# Close current window/tab
driver.close()
driver.switch_to.window(main_window)
```

## Lab Exercise
1. Automate a scenario: click button → alert appears → type in prompt → verify result
2. Switch to an iframe, fill a form inside it, submit, switch back to main content
3. Open a link in a new tab, verify the URL, close it, and return to the original tab
''')

write(S,'_16_03_04_javascript_executor.md',
fm('16_03_04','JavaScript Executor','Selenium',3,'Advanced Interactions',4,'intermediate',
   ['execute_script','execute_async_script','scroll','click','getAttribute','DOM','shadow-DOM','hidden-elements']) + '''
## execute_script

```python
# Pass arguments using `arguments[0]`, `arguments[1]`, etc.

# Scroll to element
element = driver.find_element(By.ID, "footer")
driver.execute_script("arguments[0].scrollIntoView(true);", element)

# Scroll to position
driver.execute_script("window.scrollTo(0, document.body.scrollHeight);")
driver.execute_script("window.scrollTo(0, 0);")   # scroll to top

# Click element (bypasses visibility check)
driver.execute_script("arguments[0].click();", element)

# Set value (bypasses read-only or custom inputs)
driver.execute_script("arguments[0].value = arguments[1];", element, "new value")

# Get value
val = driver.execute_script("return arguments[0].value;", element)
txt = driver.execute_script("return arguments[0].innerText;", element)

# Modify style
driver.execute_script("arguments[0].style.border = '2px solid red';", element)

# Remove attribute
driver.execute_script("arguments[0].removeAttribute('readonly');", element)
```

## execute_async_script

```python
# For async operations (AJAX, setTimeout, etc.)
result = driver.execute_async_script("""
    var callback = arguments[arguments.length - 1];
    setTimeout(function() {
        callback("done after 2s");
    }, 2000);
""")
print(result)   # "done after 2s"
```

## Shadow DOM

```python
# Access shadow root (Selenium 4)
host = driver.find_element(By.CSS_SELECTOR, "my-component")
shadow_root = driver.execute_script("return arguments[0].shadowRoot", host)
inner_el = shadow_root.find_element(By.CSS_SELECTOR, ".inner-button")
inner_el.click()
```

## Common Use Cases

```python
# Highlight element (for debugging)
def highlight(driver, element):
    driver.execute_script(
        "arguments[0].style.backgroundColor = 'yellow'; "
        "arguments[0].style.border = '2px solid red';",
        element
    )

# Get page dimensions
width  = driver.execute_script("return document.body.scrollWidth;")
height = driver.execute_script("return document.body.scrollHeight;")

# Check if element is in viewport
in_view = driver.execute_script("""
    var rect = arguments[0].getBoundingClientRect();
    return rect.top >= 0 && rect.bottom <= window.innerHeight;
""", element)
```

## Lab Exercise
1. Use JS executor to interact with a date-picker hidden behind CSS `display:none`
2. Scroll through a long page in 500px increments, scraping content at each step
3. Access and interact with a Shadow DOM component (e.g., a custom web component)
''')

write(S,'_16_03_05_file_upload_and_download.md',
fm('16_03_05','File Upload and Download','Selenium',3,'Advanced Interactions',5,'intermediate',
   ['file-upload','input-type-file','send_keys','download','chrome-prefs','robot','pyautogui','wait-for-download']) + '''
## File Upload

```python
# Standard <input type="file"> — just send_keys the absolute path
file_input = driver.find_element(By.CSS_SELECTOR, "input[type='file']")
file_input.send_keys(r"C:/path/to/my/document.pdf")

# Multiple files
file_input.send_keys(
    r"C:/files/file1.pdf" + "\n" + r"C:/files/file2.pdf"
)
# OR send separately
file_input.send_keys(r"C:/files/file1.pdf")
```

## File Download Configuration

```python
import os

download_dir = os.path.abspath("downloads")
os.makedirs(download_dir, exist_ok=True)

# Chrome — set download directory
prefs = {
    "download.default_directory": download_dir,
    "download.prompt_for_download": False,
    "download.directory_upgrade": True,
    "safebrowsing.enabled": True,
}
options = webdriver.ChromeOptions()
options.add_experimental_option("prefs", prefs)
driver = webdriver.Chrome(options=options)
```

## Wait for Download to Complete

```python
import time
import glob

def wait_for_download(directory, filename_pattern="*.pdf", timeout=30):
    """Wait until a file matching pattern appears and is complete"""
    deadline = time.time() + timeout
    while time.time() < deadline:
        files = glob.glob(os.path.join(directory, filename_pattern))
        # Filter out .crdownload (Chrome temp files)
        complete = [f for f in files if not f.endswith(".crdownload")]
        if complete:
            return complete[-1]   # return latest matching file
        time.sleep(0.5)
    raise TimeoutError(f"Download not completed in {timeout}s")

# Usage
driver.find_element(By.ID, "download-report").click()
downloaded = wait_for_download(download_dir, "report_*.xlsx")
print(f"Downloaded: {downloaded}")
```

## Handling Native OS Dialogs

```python
# For systems without hidden file inputs (uses pyautogui)
import pyautogui, time

driver.find_element(By.ID, "upload-button").click()
time.sleep(1)   # wait for native dialog to open

pyautogui.typewrite(r"C:\files\document.pdf", interval=0.05)
pyautogui.press("enter")
```

## Lab Exercise
1. Upload a PDF to a file-upload form and verify the file name appears in the UI
2. Configure Chrome to auto-download CSVs to a temp directory, trigger download, verify file
3. Handle a chunked upload form that shows a progress bar — wait until 100% complete
''')

write(S,'_16_04_01_page_object_model_pattern.md',
fm('16_04_01','Page Object Model Pattern','Selenium',4,'Test Architecture',1,'intermediate',
   ['POM','page-object','BasePage','locators','actions','separation-of-concerns','maintainability','DRY']) + '''
## Why Page Object Model?

POM separates **test logic** from **page interaction code**:

- **Without POM**: Locators scattered throughout tests → fragile
- **With POM**: Each page has one class → change locator in one place

## Structure

```
tests/
    test_login.py
    test_checkout.py
pages/
    base_page.py
    login_page.py
    home_page.py
    checkout_page.py
conftest.py
```

## Base Page

```python
# pages/base_page.py
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC

class BasePage:
    URL = ""

    def __init__(self, driver):
        self.driver = driver
        self.wait = WebDriverWait(driver, 10)

    def open(self):
        self.driver.get(self.URL)
        return self

    def find(self, locator):
        return self.wait.until(EC.visibility_of_element_located(locator))

    def click(self, locator):
        self.wait.until(EC.element_to_be_clickable(locator)).click()

    def type(self, locator, text):
        element = self.find(locator)
        element.clear()
        element.send_keys(text)

    def get_text(self, locator):
        return self.find(locator).text

    def is_displayed(self, locator):
        try:
            return self.find(locator).is_displayed()
        except Exception:
            return False
```

## Page Object

```python
# pages/login_page.py
from selenium.webdriver.common.by import By
from .base_page import BasePage

class LoginPage(BasePage):
    URL = "https://myapp.com/login"

    # Locators (as class attributes)
    USERNAME_INPUT  = (By.ID, "username")
    PASSWORD_INPUT  = (By.ID, "password")
    SUBMIT_BUTTON   = (By.CSS_SELECTOR, "button[type='submit']")
    ERROR_MESSAGE   = (By.CLASS_NAME, "error-alert")
    FORGOT_PASSWORD = (By.LINK_TEXT, "Forgot Password?")

    # Actions
    def login(self, username: str, password: str):
        self.type(self.USERNAME_INPUT, username)
        self.type(self.PASSWORD_INPUT, password)
        self.click(self.SUBMIT_BUTTON)
        return self   # fluent interface

    def get_error(self) -> str:
        return self.get_text(self.ERROR_MESSAGE)

    def has_error(self) -> bool:
        return self.is_displayed(self.ERROR_MESSAGE)
```

## Test Using POM

```python
# tests/test_login.py
import pytest
from pages.login_page import LoginPage
from pages.home_page import HomePage

def test_valid_login(driver):
    login = LoginPage(driver).open()
    login.login("user@test.com", "password123")

    home = HomePage(driver)
    assert home.is_logged_in()
    assert home.get_username() == "user@test.com"

def test_invalid_login(driver):
    login = LoginPage(driver).open()
    login.login("wrong@test.com", "wrongpass")

    assert login.has_error()
    assert "Invalid credentials" in login.get_error()
```

## Lab Exercise
1. Build POM for a 3-page e-commerce flow: Home → Product → Cart
2. Add a `navigate_to()` method to BasePage that waits for the URL to change
3. Implement a `DriverFactory` that creates Chrome/Firefox/Edge drivers by config
''')

write(S,'_16_04_02_page_factory_pattern.md',
fm('16_04_02','Page Factory Pattern','Selenium',4,'Test Architecture',2,'intermediate',
   ['page-factory','component-objects','reusable','header','footer','navigation','composition','PageComponent']) + '''
## Page Factory — Composable Components

Rather than one monolithic page object, break pages into reusable **components** (header, footer, nav bar, modal, etc.).

```python
# components/header.py
from selenium.webdriver.common.by import By
from pages.base_page import BasePage

class Header:
    LOGO        = (By.CSS_SELECTOR, ".logo")
    SEARCH_BOX  = (By.ID, "global-search")
    CART_ICON   = (By.CSS_SELECTOR, ".cart-count")
    USER_MENU   = (By.ID, "user-dropdown")

    def __init__(self, driver):
        self.driver = driver
        self._base = BasePage(driver)

    def search(self, query: str):
        self._base.type(self.SEARCH_BOX, query)
        from selenium.webdriver.common.keys import Keys
        self._base.find(self.SEARCH_BOX).send_keys(Keys.RETURN)

    def get_cart_count(self) -> int:
        text = self._base.get_text(self.CART_ICON)
        return int(text) if text.isdigit() else 0

    def open_user_menu(self):
        self._base.click(self.USER_MENU)
```

## Page that Composes Components

```python
# pages/product_listing_page.py
from .base_page import BasePage
from components.header import Header
from components.product_card import ProductCard
from selenium.webdriver.common.by import By

class ProductListingPage(BasePage):
    URL = "https://myapp.com/products"
    PRODUCT_CARDS = (By.CSS_SELECTOR, ".product-card")
    SORT_SELECT   = (By.ID, "sort-by")

    def __init__(self, driver):
        super().__init__(driver)
        self.header = Header(driver)        # composition

    def get_products(self) -> list:
        cards = self.driver.find_elements(*self.PRODUCT_CARDS)
        return [ProductCard(card) for card in cards]

    def sort_by(self, option: str):
        from selenium.webdriver.support.ui import Select
        Select(self.find(self.SORT_SELECT)).select_by_visible_text(option)
```

## ProductCard Component

```python
# components/product_card.py
from selenium.webdriver.common.by import By

class ProductCard:
    def __init__(self, element):
        self._el = element

    def get_name(self) -> str:
        return self._el.find_element(By.CSS_SELECTOR, ".product-name").text

    def get_price(self) -> float:
        text = self._el.find_element(By.CSS_SELECTOR, ".price").text
        return float(text.replace("$", "").replace(",", ""))

    def add_to_cart(self):
        self._el.find_element(By.CSS_SELECTOR, ".add-to-cart").click()
```

## Lab Exercise
1. Create Header, Footer, and Modal components shared across 3 page objects
2. Write a test that uses Header.search() and verifies product listing updates
3. Implement a `DataTable` component that wraps `<table>` and provides `get_row_by_column(col, val)`
''')

write(S,'_16_04_03_base_page_and_utilities.md',
fm('16_04_03','Base Page and Utilities','Selenium',4,'Test Architecture',3,'intermediate',
   ['BasePage','utility','screenshot-on-failure','scroll','highlight','config','DriverFactory','conftest']) + '''
## Enhanced Base Page

```python
# pages/base_page.py
import time, os
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from selenium.common.exceptions import TimeoutException, NoSuchElementException

class BasePage:
    DEFAULT_TIMEOUT = 10

    def __init__(self, driver):
        self.driver = driver
        self.wait = WebDriverWait(driver, self.DEFAULT_TIMEOUT)

    def open(self, url: str = None):
        target = url or getattr(self, "URL", "")
        self.driver.get(target)
        return self

    # ── Finding ──────────────────────────────────────────────
    def find(self, locator, timeout=None):
        w = WebDriverWait(self.driver, timeout or self.DEFAULT_TIMEOUT)
        return w.until(EC.visibility_of_element_located(locator))

    def find_all(self, locator):
        return self.driver.find_elements(*locator)

    def is_present(self, locator, timeout=3) -> bool:
        try:
            WebDriverWait(self.driver, timeout).until(
                EC.presence_of_element_located(locator)
            )
            return True
        except TimeoutException:
            return False

    # ── Interacting ───────────────────────────────────────────
    def click(self, locator):
        self.wait.until(EC.element_to_be_clickable(locator)).click()

    def type(self, locator, text: str):
        el = self.find(locator)
        el.clear()
        el.send_keys(text)

    def select_option(self, locator, text: str):
        from selenium.webdriver.support.ui import Select
        Select(self.find(locator)).select_by_visible_text(text)

    # ── Reading ───────────────────────────────────────────────
    def get_text(self, locator) -> str:
        return self.find(locator).text

    def get_attr(self, locator, attribute: str) -> str:
        return self.find(locator).get_attribute(attribute)

    # ── Utilities ─────────────────────────────────────────────
    def scroll_to(self, locator):
        el = self.find(locator)
        self.driver.execute_script("arguments[0].scrollIntoView(true);", el)

    def scroll_by(self, px: int):
        self.driver.execute_script(f"window.scrollBy(0, {px});")

    def highlight(self, locator, color="yellow"):
        el = self.find(locator)
        self.driver.execute_script(
            f"arguments[0].style.backgroundColor='{color}'; "
            f"arguments[0].style.border='2px solid red';", el
        )

    def screenshot(self, name: str = None):
        fname = name or f"screenshot_{int(time.time())}.png"
        self.driver.save_screenshot(fname)
        return fname
```

## conftest.py — Pytest Fixtures

```python
# conftest.py
import pytest
from selenium import webdriver
from selenium.webdriver.chrome.options import Options

@pytest.fixture(scope="session")
def driver():
    opts = Options()
    opts.add_argument("--start-maximized")
    # opts.add_argument("--headless")
    drv = webdriver.Chrome(options=opts)
    drv.implicitly_wait(0)   # use explicit waits only
    yield drv
    drv.quit()

@pytest.fixture(autouse=True)
def screenshot_on_failure(driver, request):
    yield
    if request.node.rep_call.failed:
        test_name = request.node.name
        driver.save_screenshot(f"failures/{test_name}.png")

@pytest.hookimpl(hookwrapper=True)
def pytest_runtest_makereport(item, call):
    outcome = yield
    rep = outcome.get_result()
    setattr(item, "rep_" + rep.when, rep)
```

## Lab Exercise
1. Add `wait_for_url_contains(partial_url)` and `wait_for_title_contains(text)` to BasePage
2. Configure `conftest.py` to take a screenshot on EVERY failure automatically
3. Add `DriverFactory.create(browser: str, headless: bool)` that supports Chrome, Firefox, Edge
''')

write(S,'_16_05_01_pytest_with_selenium.md',
fm('16_05_01','Pytest with Selenium','Selenium',5,'Testing Framework Integration',1,'intermediate',
   ['pytest','fixture','conftest','mark','parametrize','assert','setup','teardown','allure','scope']) + '''
## Test Structure

```python
# tests/test_login.py
import pytest
from pages.login_page import LoginPage
from pages.home_page import HomePage

class TestLogin:
    def test_valid_credentials(self, driver, base_url):
        page = LoginPage(driver)
        page.open(base_url + "/login")
        page.login("admin@test.com", "password123")
        assert HomePage(driver).is_logged_in()

    def test_empty_username(self, driver, base_url):
        page = LoginPage(driver)
        page.open(base_url + "/login")
        page.login("", "password123")
        assert page.has_error()
        assert "Username is required" in page.get_error()

    @pytest.mark.skip(reason="Feature in development")
    def test_sso_login(self, driver):
        ...
```

## Fixtures Hierarchy

```python
# conftest.py
import pytest
from selenium import webdriver

@pytest.fixture(scope="session")           # once per session
def driver():
    drv = webdriver.Chrome()
    yield drv
    drv.quit()

@pytest.fixture(scope="module")            # once per module
def authenticated_driver(driver, base_url):
    from pages.login_page import LoginPage
    LoginPage(driver).open(base_url + "/login").login("user", "pass")
    yield driver
    driver.delete_all_cookies()

@pytest.fixture(scope="function")          # default — before each test
def fresh_page(driver):
    driver.delete_all_cookies()
    driver.get("about:blank")
    yield

@pytest.fixture
def base_url():
    return "https://staging.myapp.com"
```

## Markers

```python
# pytest.ini or pyproject.toml
[pytest]
markers =
    smoke: Quick sanity check tests
    regression: Full regression suite
    slow: Tests taking >30 seconds

# Usage
@pytest.mark.smoke
def test_homepage_loads(driver): ...

@pytest.mark.regression
@pytest.mark.slow
def test_full_checkout(driver): ...

# Run only smoke tests
# pytest -m smoke
```

## Allure Reporting

```bash
pip install allure-pytest

pytest --alluredir=allure-results tests/
allure serve allure-results
```

```python
import allure

@allure.title("User can login with valid credentials")
@allure.description("Validates the happy path for user authentication")
@allure.feature("Authentication")
@allure.severity(allure.severity_level.CRITICAL)
def test_login(driver):
    with allure.step("Open login page"):
        driver.get("https://myapp.com/login")
    with allure.step("Enter credentials"):
        driver.find_element(By.ID, "username").send_keys("user@test.com")
    with allure.step("Submit form"):
        driver.find_element(By.ID, "submit").click()
    allure.attach(driver.get_screenshot_as_png(),
                  name="screenshot", attachment_type=allure.attachment_type.PNG)
```

## Lab Exercise
1. Create a full test suite for a registration flow with 5 test cases using class-based tests
2. Set up `scope="session"` driver with login in `autouse` fixture
3. Add Allure report generation with screenshots on each test step
''')

write(S,'_16_05_02_test_configuration_and_reporting.md',
fm('16_05_02','Test Configuration and Reporting','Selenium',5,'Testing Framework Integration',2,'intermediate',
   ['pytest.ini','conftest','html-report','allure','extent-reports','environment','parametrize','parallel','xdist']) + '''
## pytest.ini / pyproject.toml

```ini
# pytest.ini
[pytest]
testpaths = tests
python_files = test_*.py
python_classes = Test*
python_functions = test_*
addopts =
    -v
    --tb=short
    --screenshot=on
    --html=reports/report.html
    --self-contained-html
markers =
    smoke: Smoke tests
    regression: Full regression
    slow: > 30 seconds
```

## HTML Report

```bash
pip install pytest-html

pytest --html=reports/report.html --self-contained-html tests/
```

## Parallel Execution with pytest-xdist

```bash
pip install pytest-xdist

pytest -n 4 tests/          # 4 parallel workers
pytest -n auto tests/       # auto-detect CPU count
```

```python
# conftest.py — unique driver per worker
@pytest.fixture(scope="function")
def driver(tmp_path):
    import uuid
    opts = webdriver.ChromeOptions()
    opts.add_argument(f"--user-data-dir=/tmp/chrome_{uuid.uuid4()}")
    drv = webdriver.Chrome(options=opts)
    yield drv
    drv.quit()
```

## Environment Configuration

```python
# config.py
import os

class Config:
    BASE_URL     = os.getenv("BASE_URL", "https://staging.myapp.com")
    USERNAME     = os.getenv("TEST_USER", "testuser@example.com")
    PASSWORD     = os.getenv("TEST_PASS", "Test@1234")
    BROWSER      = os.getenv("BROWSER", "chrome")
    HEADLESS     = os.getenv("HEADLESS", "false").lower() == "true"
    TIMEOUT      = int(os.getenv("TIMEOUT", "10"))

config = Config()
```

## Lab Exercise
1. Set up pytest.ini with custom markers, test discovery, and default HTML report generation
2. Run the same test suite in parallel with `-n 4` and verify isolation
3. Read base URL and credentials from environment variables using `os.getenv`
''')

write(S,'_16_05_03_data_driven_testing.md',
fm('16_05_03','Data Driven Testing','Selenium',5,'Testing Framework Integration',3,'intermediate',
   ['parametrize','csv','excel','json','data-driven','DDT','test-data','openpyxl','faker']) + '''
## pytest.mark.parametrize

```python
import pytest

LOGIN_DATA = [
    ("admin@test.com",   "Admin@123",    True,  "Dashboard"),
    ("editor@test.com",  "Editor@123",   True,  "Editor Panel"),
    ("wrong@test.com",   "wrongpass",    False, ""),
    ("admin@test.com",   "",             False, ""),
]

@pytest.mark.parametrize("username,password,expected_success,expected_title",
                         LOGIN_DATA, ids=["admin","editor","wrong-creds","empty-pass"])
def test_login_scenarios(driver, username, password, expected_success, expected_title):
    page = LoginPage(driver).open()
    page.login(username, password)

    if expected_success:
        assert HomePage(driver).get_title() == expected_title
    else:
        assert page.has_error()
```

## Reading Test Data from CSV

```python
import csv

def load_csv(filepath):
    with open(filepath, newline="", encoding="utf-8") as f:
        reader = csv.DictReader(f)
        return list(reader)

test_data = load_csv("test_data/login_data.csv")

@pytest.mark.parametrize("data", test_data,
    ids=[d["test_id"] for d in test_data])
def test_login_csv(driver, data):
    page = LoginPage(driver).open()
    page.login(data["username"], data["password"])
    if data["expected"] == "success":
        assert HomePage(driver).is_logged_in()
    else:
        assert page.has_error()
```

## Reading from Excel

```python
import openpyxl

def load_excel(filepath, sheet_name="Sheet1"):
    wb = openpyxl.load_workbook(filepath)
    ws = wb[sheet_name]
    headers = [cell.value for cell in ws[1]]
    return [
        dict(zip(headers, [cell.value for cell in row]))
        for row in ws.iter_rows(min_row=2)
    ]
```

## Generating Data with Faker

```python
from faker import Faker
fake = Faker()

def generate_user():
    return {
        "name": fake.name(),
        "email": fake.email(),
        "phone": fake.phone_number(),
        "address": fake.address(),
    }

@pytest.mark.parametrize("user", [generate_user() for _ in range(5)])
def test_register_new_user(driver, user):
    page = RegistrationPage(driver).open()
    page.fill_form(**user)
    page.submit()
    assert page.success_message_displayed()
```

## Lab Exercise
1. Load 20 product searches from a CSV and verify each search returns results
2. Parametrize a checkout test with different shipping addresses from an Excel file
3. Use Faker to generate and test 10 unique user registrations
''')

write(S,'_16_06_01_headless_browser_testing.md',
fm('16_06_01','Headless Browser Testing','Selenium',6,'Advanced and CI',1,'intermediate',
   ['headless','--headless','Chrome-headless','Firefox-headless','screenshots','performance','CI']) + '''
## What is Headless?

Headless mode runs the browser **without a visible window**. Essential for:
- CI/CD pipelines (no display available)
- Faster test execution
- Server-side scraping

## Chrome Headless

```python
from selenium.webdriver.chrome.options import Options

opts = Options()
opts.add_argument("--headless=new")          # new headless mode (Chrome 112+)
opts.add_argument("--window-size=1920,1080") # set viewport
opts.add_argument("--disable-gpu")           # needed on some systems
opts.add_argument("--no-sandbox")            # required in Docker
opts.add_argument("--disable-dev-shm-usage") # prevent /dev/shm issues in Docker

driver = webdriver.Chrome(options=opts)
driver.get("https://example.com")
driver.save_screenshot("headless_screenshot.png")
driver.quit()
```

## Firefox Headless

```python
from selenium.webdriver.firefox.options import Options as FirefoxOptions

opts = FirefoxOptions()
opts.add_argument("--headless")
driver = webdriver.Firefox(options=opts)
```

## Headless with Virtual Display (Linux)

```bash
# Install Xvfb (virtual frame buffer)
sudo apt-get install xvfb
Xvfb :99 -screen 0 1920x1080x24 &
export DISPLAY=:99

# OR use pyvirtualdisplay in Python
pip install pyvirtualdisplay
```

```python
from pyvirtualdisplay import Display

display = Display(visible=0, size=(1920, 1080))
display.start()

driver = webdriver.Chrome()   # now uses virtual display
# run tests...
driver.quit()
display.stop()
```

## Lab Exercise
1. Configure headless Chrome with 1920x1080 viewport and run 5 tests
2. Set up GitHub Actions workflow that runs headless Selenium tests on push
3. Compare execution time: headed vs headless for a 10-test suite
''')

write(S,'_16_06_02_selenium_grid.md',
fm('16_06_02','Selenium Grid','Selenium',6,'Advanced and CI',2,'advanced',
   ['selenium-grid','hub','node','capabilities','remote-webdriver','parallel','docker-selenium','browser-farm']) + '''
## Selenium Grid Architecture

```
              ┌─────────────────────────┐
              │        Grid Hub          │
              │  (distributes sessions)  │
              └────────────┬────────────┘
                           │
           ┌───────────────┼───────────────┐
           │               │               │
    ┌──────┴──────┐ ┌──────┴──────┐ ┌──────┴──────┐
    │  Node: Win  │ │  Node: Mac  │ │  Node: Linux│
    │  Chrome,IE  │ │  Safari,FF  │ │  Chrome,FF  │
    └─────────────┘ └─────────────┘ └─────────────┘
```

## Standalone Grid (Single Node)

```bash
# Download selenium-server-4.x.jar from selenium.dev/downloads

# Start standalone (hub+node in one)
java -jar selenium-server-4.x.jar standalone

# Start hub
java -jar selenium-server-4.x.jar hub

# Start node
java -jar selenium-server-4.x.jar node --hub http://localhost:4444
```

## Remote WebDriver

```python
from selenium import webdriver
from selenium.webdriver.common.desired_capabilities import DesiredCapabilities

# Connect to Grid hub
options = webdriver.ChromeOptions()
options.set_capability("browserVersion", "latest")
options.set_capability("platformName", "Windows 10")

driver = webdriver.Remote(
    command_executor="http://localhost:4444/wd/hub",
    options=options,
)

driver.get("https://example.com")
print(driver.title)
driver.quit()
```

## Docker Selenium Grid

```yaml
# docker-compose.yml
version: "3.8"
services:
  selenium-hub:
    image: selenium/hub:4
    ports:
      - "4444:4444"

  chrome:
    image: selenium/node-chrome:4
    depends_on: [selenium-hub]
    environment:
      - SE_EVENT_BUS_HOST=selenium-hub
    volumes:
      - /dev/shm:/dev/shm

  firefox:
    image: selenium/node-firefox:4
    depends_on: [selenium-hub]
    environment:
      - SE_EVENT_BUS_HOST=selenium-hub
```

```bash
docker-compose up -d
pytest tests/ -n 4  # parallel across grid nodes
```

## Lab Exercise
1. Start a Selenium Grid with Docker Compose (hub + Chrome node)
2. Run 5 tests simultaneously on the Grid using `webdriver.Remote`
3. Add a Firefox node and run tests cross-browser in parallel
''')

write(S,'_16_06_03_ci_cd_integration.md',
fm('16_06_03','CI/CD Integration','Selenium',6,'Advanced and CI',3,'intermediate',
   ['github-actions','jenkins','CI','CD','headless','artifacts','allure','docker','environment-variables','test-report']) + '''
## GitHub Actions Workflow

```yaml
# .github/workflows/selenium-tests.yml
name: Selenium UI Tests

on:
  push:
    branches: [main, develop]
  pull_request:
    branches: [main]

jobs:
  ui-tests:
    runs-on: ubuntu-latest

    steps:
      - uses: actions/checkout@v4

      - name: Set up Python
        uses: actions/setup-python@v5
        with:
          python-version: "3.12"

      - name: Install dependencies
        run: |
          pip install -r requirements.txt

      - name: Run Selenium Tests
        env:
          BASE_URL: ${{ secrets.STAGING_URL }}
          TEST_USER: ${{ secrets.TEST_USER }}
          TEST_PASS: ${{ secrets.TEST_PASS }}
        run: |
          pytest tests/ \
            --headless \
            --html=reports/report.html \
            --self-contained-html \
            -v

      - name: Upload test report
        if: always()
        uses: actions/upload-artifact@v4
        with:
          name: test-report
          path: reports/

      - name: Upload failure screenshots
        if: failure()
        uses: actions/upload-artifact@v4
        with:
          name: failure-screenshots
          path: failures/
```

## Jenkins Pipeline (Declarative)

```groovy
pipeline {
    agent any

    stages {
        stage('Install') {
            steps {
                sh 'pip install -r requirements.txt'
            }
        }
        stage('Test') {
            steps {
                sh """
                    pytest tests/ \
                      --html=report.html \
                      --self-contained-html \
                      -n 4
                """
            }
        }
    }

    post {
        always {
            publishHTML(target: [
                reportDir: '.',
                reportFiles: 'report.html',
                reportName: 'Selenium Test Report'
            ])
        }
    }
}
```

## Lab Exercise
1. Create a GitHub Actions workflow that runs tests on every PR targeting `main`
2. Archive screenshots from failed tests as workflow artifacts
3. Add a Slack notification step that posts the pass/fail summary on completion
''')

write(S,'_16_06_04_screenshot_and_visual_testing.md',
fm('16_06_04','Screenshot and Visual Testing','Selenium',6,'Advanced and CI',4,'intermediate',
   ['screenshot','visual-regression','pixel-comparison','Pillow','Percy','Applitools','baseline','diff']) + '''
## Screenshot Capture

```python
import os, time

def take_screenshot(driver, name, folder="screenshots"):
    os.makedirs(folder, exist_ok=True)
    path = os.path.join(folder, f"{name}_{int(time.time())}.png")
    driver.save_screenshot(path)
    return path

# Element-level screenshot
element = driver.find_element(By.ID, "product-card")
element.screenshot("product_card.png")

# Full-page screenshot (Selenium 4.15+)
from selenium.webdriver.common.print_page_options import PrintOptions
print_opts = PrintOptions()
print_opts.orientation = "portrait"
pdf = driver.print_page(print_opts)   # returns base64 PDF
```

## Basic Image Comparison with Pillow

```python
from PIL import Image, ImageChops
import math, operator, functools

def images_are_equal(img1_path, img2_path, threshold=0.01):
    img1 = Image.open(img1_path).convert("RGB")
    img2 = Image.open(img2_path).convert("RGB")

    if img1.size != img2.size:
        return False

    diff = ImageChops.difference(img1, img2)
    pixels = list(diff.getdata())
    total_diff = sum(sum(p) for p in pixels)
    max_diff = 255 * 3 * img1.size[0] * img1.size[1]
    diff_ratio = total_diff / max_diff

    return diff_ratio <= threshold

# Baseline comparison
baseline = "baselines/home_page.png"
current = take_screenshot(driver, "home_page_current")

if os.path.exists(baseline):
    if not images_are_equal(baseline, current):
        print("VISUAL REGRESSION DETECTED!")
else:
    import shutil
    shutil.copy(current, baseline)
    print("Baseline created")
```

## Lab Exercise
1. Build a visual regression framework: capture baselines on first run, compare on subsequent runs
2. Highlight pixel differences between two screenshots by drawing red borders around diff areas
3. Integrate Percy.io visual testing into a pytest test (using Percy SDK)
''')

write(S,'_16_06_05_capstone_ecommerce_automation.md',
fm('16_06_05','Capstone E-Commerce Automation','Selenium',6,'Advanced and CI',5,'advanced',
   ['capstone','e-commerce','end-to-end','POM','pytest','data-driven','CI','report','screenshot']) + '''
## Project Overview

Build a complete automated test suite for an e-commerce website (e.g., `https://automationexercise.com`) covering:

1. User registration and login
2. Product search and filtering
3. Add to cart and checkout
4. Order history verification

## Project Structure

```
ecommerce_tests/
    pages/
        base_page.py
        login_page.py
        register_page.py
        product_listing_page.py
        product_detail_page.py
        cart_page.py
        checkout_page.py
        order_confirmation_page.py
    components/
        header.py
        footer.py
        product_card.py
    tests/
        test_auth.py
        test_search.py
        test_cart.py
        test_checkout.py
    test_data/
        users.csv
        products.json
    conftest.py
    pytest.ini
    requirements.txt
```

## Key Test Scenarios

```python
# tests/test_checkout.py
import pytest
from faker import Faker
from pages.login_page import LoginPage
from pages.product_listing_page import ProductListingPage
from pages.cart_page import CartPage
from pages.checkout_page import CheckoutPage

fake = Faker()

class TestCheckout:
    @pytest.mark.smoke
    def test_guest_checkout(self, driver):
        """Guest user can complete checkout with valid details"""
        # 1. Add product to cart
        ProductListingPage(driver).open() \
            .search("blue dress") \
            .get_first_product().add_to_cart()

        # 2. Go to cart
        cart = CartPage(driver).open()
        assert cart.item_count() == 1

        # 3. Proceed to checkout
        checkout = cart.proceed_to_checkout()
        checkout.fill_shipping(
            name=fake.name(), email=fake.email(),
            address=fake.address(), zip=fake.zipcode()
        )
        checkout.pay_with_card(
            number="4111111111111111", expiry="12/26", cvv="123"
        )

        # 4. Verify confirmation
        confirmation = checkout.place_order()
        assert confirmation.order_placed()
        assert confirmation.get_order_id() is not None

    @pytest.mark.regression
    @pytest.mark.parametrize("product", ["t-shirt", "jeans", "jacket"])
    def test_add_multiple_products(self, driver, authenticated_user, product):
        page = ProductListingPage(driver).open()
        page.search(product).get_first_product().add_to_cart()
        assert CartPage(driver).open().item_count() >= 1
```

## Deliverables

| Artifact | Description |
|---|---|
| `conftest.py` | Session driver, authenticated fixture, screenshot on fail |
| `pytest.ini` | Markers, options, HTML report config |
| `pages/` | Full POM with BasePage, 8 page classes |
| `components/` | Header, Cart icon, Product card |
| `tests/` | 20+ test cases across 4 modules |
| `reports/` | HTML + Allure report |
| `.github/workflows/` | CI pipeline on push |

## Lab Exercise
1. Implement all 8 page objects with proper locators and action methods
2. Write 20 test cases with data-driven login, search, and checkout tests
3. Set up CI pipeline to run on push, upload report, and notify on failure
''')

# ═══════════════════════════════════════════════════════════════
# JAVA — 21 lessons
# ═══════════════════════════════════════════════════════════════
print()
print('='*60)
print('JAVA — 21 lessons')
print('='*60)
J = '_08_java'

java_lessons = {
'_08_01_01_java_overview_and_setup.md': ('08_01_01','Java Overview and Setup','Java',1,'Java Fundamentals',1,'beginner',['JDK','JVM','JRE','Maven','Gradle','IntelliJ','javac','java','WORA','bytecode'],'''
## What is Java?

Java is a **statically typed, object-oriented** language developed by Sun Microsystems (1995), now maintained by Oracle. Key principle: **Write Once Run Anywhere (WORA)** — code compiles to bytecode executed by the JVM on any platform.

### JVM / JRE / JDK

| Component | Contents | Who Needs It |
|---|---|---|
| **JVM** | Java Virtual Machine (executes bytecode) | Runtime |
| **JRE** | JVM + standard libraries | Running Java apps |
| **JDK** | JRE + compiler (javac) + tools | Developing Java |

## Installation

```bash
# Windows — via Winget
winget install Microsoft.OpenJDK.21

# Ubuntu
sudo apt install openjdk-21-jdk

# Verify
java --version
javac --version
```

## Hello World

```java
// HelloWorld.java
public class HelloWorld {
    public static void main(String[] args) {
        System.out.println("Hello, World!");
        System.out.printf("Java %s%n", System.getProperty("java.version"));
    }
}
```

```bash
javac HelloWorld.java   # compiles to HelloWorld.class (bytecode)
java HelloWorld         # runs on JVM
```

## Build Tools

```xml
<!-- Maven pom.xml -->
<project>
  <groupId>com.example</groupId>
  <artifactId>myapp</artifactId>
  <version>1.0.0</version>
  <properties>
    <maven.compiler.source>21</maven.compiler.source>
    <maven.compiler.target>21</maven.compiler.target>
  </properties>
  <dependencies>
    <dependency>
      <groupId>com.google.code.gson</groupId>
      <artifactId>gson</artifactId>
      <version>2.10.1</version>
    </dependency>
  </dependencies>
</project>
```

```bash
mvn compile
mvn test
mvn package   # builds JAR
```

## Lab Exercise
1. Install JDK 21, verify with `java --version`
2. Write and compile a program that prints system info: OS, Java version, available processors
3. Create a Maven project structure and add a Gson dependency
'''),

'_08_01_02_data_types_variables_operators.md': ('08_01_02','Data Types Variables and Operators','Java',1,'Java Fundamentals',2,'beginner',['primitive','int','long','double','boolean','char','String','var','final','operators','casting'],'''
## Primitive Types

```java
// Integer types
byte   b = 127;              // 8-bit  (-128 to 127)
short  s = 32767;            // 16-bit
int    i = 2_147_483_647;    // 32-bit (default integer)
long   l = 9_223_372_036_854_775_807L;  // 64-bit (suffix L)

// Floating point
float  f = 3.14f;            // 32-bit (suffix f)
double d = 3.141592653589793; // 64-bit (default decimal)

// Other
boolean flag = true;
char    c    = 'A';          // 16-bit Unicode character (UTF-16)

// Literals
int hex = 0xFF;              // 255
int bin = 0b1010;            // 10
long big = 1_000_000L;       // underscore for readability
```

## Reference Types and Strings

```java
// String — immutable, interned
String name = "Raja";
String greeting = "Hello, " + name + "!";
String multiline = """
        Line 1
        Line 2
        """;  // Text block (Java 15+)

// String methods
name.length()           // 4
name.toUpperCase()      // "RAJA"
name.charAt(0)          // 'R'
name.substring(1, 3)    // "aj"
name.contains("aj")     // true
name.replace("a", "A")  // "RAjA"
name.strip()            // trim (Unicode-aware)
String.format("Name: %s, Age: %d", name, 28)
```

## Type Inference with `var`

```java
var message = "Hello";       // String
var count   = 42;            // int
var prices  = new ArrayList<Double>();

// Works in for-each
for (var item : prices) {
    System.out.println(item);
}
```

## Constants

```java
final int MAX_SIZE = 100;
final double PI = 3.141592653589793;
// MAX_SIZE = 200; // CompileError: cannot assign final
```

## Type Casting

```java
// Widening (implicit — safe)
int i = 100;
long l = i;       // int → long
double d = l;     // long → double

// Narrowing (explicit — may lose data)
double price = 9.99;
int truncated = (int) price;   // 9

// String conversions
int n = Integer.parseInt("42");
double pi = Double.parseDouble("3.14");
String s = String.valueOf(42);    // "42"
String.valueOf(true)              // "true"
```

## Operators

```java
// Arithmetic: + - * / % (integer division truncates)
10 / 3      // 3 (not 3.33!)
10.0 / 3    // 3.333...
10 % 3      // 1

// Comparison: == != < > <= >=
// Logical: && || !
// Bitwise: & | ^ ~ << >> >>>
// Ternary
String result = score >= 60 ? "Pass" : "Fail";

// String concatenation
"Hello" + 42       // "Hello42"
"Sum: " + (1 + 2)  // "Sum: 3"
```

## Lab Exercise
1. Calculate the area and circumference of a circle using `Math.PI`
2. Demonstrate widening and narrowing casts with a temperature converter (Celsius ↔ Fahrenheit)
3. Use a text block to format a multi-line JSON string without escape characters
'''),

'_08_01_03_control_flow.md': ('08_01_03','Control Flow','Java',1,'Java Fundamentals',3,'beginner',['if','switch','for','while','do-while','break','continue','enhanced-for','switch-expression','pattern-matching'],'''
## Conditional Statements

```java
// if / else if / else
int score = 82;
if (score >= 90) {
    System.out.println("A");
} else if (score >= 75) {
    System.out.println("B");
} else if (score >= 60) {
    System.out.println("C");
} else {
    System.out.println("F");
}

// Traditional switch
switch (day) {
    case "MON": case "TUE": case "WED": case "THU": case "FRI":
        System.out.println("Weekday");
        break;
    case "SAT": case "SUN":
        System.out.println("Weekend");
        break;
    default:
        System.out.println("Unknown");
}

// Switch expression (Java 14+)
String type = switch (day) {
    case "MON", "TUE", "WED", "THU", "FRI" -> "Weekday";
    case "SAT", "SUN" -> "Weekend";
    default -> throw new IllegalArgumentException("Unknown: " + day);
};
```

## Loops

```java
// for loop
for (int i = 0; i < 10; i++) {
    System.out.print(i + " ");
}

// while
int n = 1;
while (n <= 100) {
    n *= 2;
}

// do-while (runs at least once)
do {
    input = scanner.nextLine();
} while (input.isEmpty());

// Enhanced for (for-each)
int[] numbers = {1, 2, 3, 4, 5};
for (int num : numbers) {
    System.out.println(num);
}

// break and continue
for (int i = 0; i < 10; i++) {
    if (i == 5) break;      // exit loop
    if (i % 2 == 0) continue;  // skip even
    System.out.print(i);
}

// Labeled break (for nested loops)
outer:
for (int i = 0; i < 5; i++) {
    for (int j = 0; j < 5; j++) {
        if (i == 2 && j == 2) break outer;
        System.out.print(i + "" + j + " ");
    }
}
```

## Pattern Matching (Java 16+)

```java
Object obj = "Hello";

// instanceof with pattern variable
if (obj instanceof String s) {
    System.out.println(s.toUpperCase());  // s is a String here
}

// Pattern matching in switch (Java 21)
String describe(Object o) {
    return switch (o) {
        case Integer i -> "int: " + i;
        case String s  -> "string: " + s;
        case Double d  -> "double: " + d;
        case null      -> "null";
        default        -> "other: " + o;
    };
}
```

## Lab Exercise
1. Print a multiplication table using nested for loops
2. Build a number guessing game using `while` loop and `Scanner`
3. Rewrite a 5-case switch statement using the switch expression arrow syntax
'''),

'_08_01_04_arrays_and_strings.md': ('08_01_04','Arrays and Strings','Java',1,'Java Fundamentals',4,'beginner',['array','multi-dimensional','Arrays','String','StringBuilder','StringJoiner','String-methods','charAt','split'],'''
## Arrays

```java
// Declaration and initialization
int[] nums = {1, 2, 3, 4, 5};
String[] names = new String[3];
names[0] = "Alice"; names[1] = "Bob"; names[2] = "Charlie";

// Accessing
System.out.println(nums.length);   // 5
System.out.println(nums[0]);       // 1
System.out.println(nums[nums.length - 1]);  // 5

// 2D array
int[][] matrix = {
    {1, 2, 3},
    {4, 5, 6},
    {7, 8, 9}
};
System.out.println(matrix[1][2]);  // 6

// java.util.Arrays utility
import java.util.Arrays;

Arrays.sort(nums);
int idx = Arrays.binarySearch(nums, 3);
int[] copy = Arrays.copyOf(nums, 10);       // pad with 0s
int[] range = Arrays.copyOfRange(nums, 1, 4); // [2,3,4]
System.out.println(Arrays.toString(nums));   // [1, 2, 3, 4, 5]
Arrays.fill(nums, 0);
```

## String Methods

```java
String s = "  Hello, World!  ";

s.length()                   // 17
s.trim()                     // "Hello, World!"
s.strip()                    // same (Unicode-aware)
s.toLowerCase()
s.toUpperCase()
s.charAt(7)                  // 'W'
s.indexOf("World")           // 9
s.lastIndexOf('l')           // 12
s.substring(2, 7)            // "Hello"
s.replace("World", "Java")
s.contains("Hello")          // true
s.startsWith("  H")         // true
s.endsWith("  ")             // true
s.split(", ")                // ["  Hello", "World!  "]
s.isEmpty()                  // false
s.isBlank()                  // false
String.join(", ", "a","b","c")  // "a, b, c"
```

## StringBuilder (Mutable String)

```java
// String concatenation in loop = O(n²) — use StringBuilder instead
StringBuilder sb = new StringBuilder();
for (int i = 1; i <= 5; i++) {
    sb.append(i).append(", ");  // chaining
}
sb.deleteCharAt(sb.length() - 1);  // remove last comma
sb.insert(0, "[").append("]");
String result = sb.toString();    // "[1, 2, 3, 4, 5]"

// Common methods
sb.reverse()
sb.replace(start, end, str)
sb.delete(start, end)
sb.length()
```

## StringJoiner and String.format

```java
import java.util.StringJoiner;

StringJoiner sj = new StringJoiner(", ", "[", "]");
for (String name : names) sj.add(name);
System.out.println(sj);  // [Alice, Bob, Charlie]

// Formatted strings
String.format("Name: %-15s Age: %3d", "Raja", 28)
// "Name: Raja            Age:  28"
```

## Lab Exercise
1. Implement bubble sort on an integer array, verify with `Arrays.sort`
2. Count character frequencies in a string using an array of 26 ints
3. Reverse words in a sentence using `split()` and `StringBuilder`
'''),

'_08_01_05_methods_and_varargs.md': ('08_01_05','Methods and Varargs','Java',1,'Java Fundamentals',5,'beginner',['method','return','overloading','varargs','static','recursion','Math'],'''
## Method Syntax

```java
// access-modifier return-type methodName(params) { body }
public static double calculateBMI(double weight, double height) {
    if (height == 0) throw new IllegalArgumentException("Height cannot be 0");
    return weight / (height * height);
}

// Calling
double bmi = calculateBMI(70.0, 1.75);
System.out.printf("BMI: %.2f%n", bmi);
```

## Method Overloading

```java
public static int add(int a, int b)          { return a + b; }
public static double add(double a, double b) { return a + b; }
public static int add(int a, int b, int c)   { return a + b + c; }

add(1, 2)        // int version
add(1.0, 2.5)    // double version
add(1, 2, 3)     // three-arg version
```

## Varargs

```java
public static int sum(int... numbers) {
    int total = 0;
    for (int n : numbers) total += n;
    return total;
}

sum()            // 0
sum(1, 2, 3)     // 6
sum(1, 2, 3, 4, 5) // 15

// Varargs + other params
public static String format(String template, Object... args) {
    return String.format(template, args);
}
```

## Recursion

```java
public static long factorial(int n) {
    if (n <= 1) return 1;       // base case
    return n * factorial(n - 1); // recursive case
}

public static int fibonacci(int n) {
    if (n <= 1) return n;
    return fibonacci(n-1) + fibonacci(n-2);
}

// Tail-recursive with accumulator (optimised)
public static long factorial(int n, long acc) {
    if (n <= 1) return acc;
    return factorial(n - 1, n * acc);
}
```

## Math Class

```java
Math.abs(-5)          // 5
Math.pow(2, 10)       // 1024.0
Math.sqrt(144)        // 12.0
Math.cbrt(27)         // 3.0
Math.max(3, 7)        // 7
Math.min(3, 7)        // 3
Math.round(3.7)       // 4L
Math.floor(3.9)       // 3.0
Math.ceil(3.1)        // 4.0
Math.random()         // [0.0, 1.0)
Math.log(Math.E)      // 1.0
Math.PI               // 3.14159...
```

## Lab Exercise
1. Write overloaded `area()` methods for circle, rectangle, and triangle
2. Implement `quickSort(int[] arr, int low, int high)` recursively
3. Write a varargs `max(double... values)` that returns the maximum value
'''),

'_08_02_01_classes_and_objects.md': ('08_02_01','Classes and Objects','Java',2,'Object-Oriented Programming',1,'intermediate',['class','object','constructor','this','new','getter','setter','toString','equals','hashCode'],'''
## Class Structure

```java
public class BankAccount {
    // Fields (instance variables)
    private final String accountId;
    private String owner;
    private double balance;
    private static int accountCount = 0;  // class variable

    // Constructor
    public BankAccount(String owner, double initialBalance) {
        this.accountId = "ACC" + (++accountCount);
        this.owner = owner;
        this.balance = initialBalance;
    }

    // Methods
    public void deposit(double amount) {
        if (amount <= 0) throw new IllegalArgumentException("Amount must be positive");
        this.balance += amount;
    }

    public double withdraw(double amount) {
        if (amount > balance) throw new IllegalStateException("Insufficient funds");
        this.balance -= amount;
        return amount;
    }

    // Getters / Setters
    public double getBalance() { return balance; }
    public String getOwner()   { return owner; }
    public void setOwner(String owner) { this.owner = owner; }

    // toString, equals, hashCode
    @Override
    public String toString() {
        return String.format("BankAccount[id=%s, owner=%s, balance=%.2f]",
                             accountId, owner, balance);
    }

    @Override
    public boolean equals(Object o) {
        if (this == o) return true;
        if (!(o instanceof BankAccount other)) return false;
        return accountId.equals(other.accountId);
    }

    @Override
    public int hashCode() { return accountId.hashCode(); }
}
```

## Creating Objects

```java
BankAccount acc = new BankAccount("Raja", 1000.0);
acc.deposit(500.0);
System.out.println(acc.getBalance());   // 1500.0
System.out.println(acc);               // BankAccount[id=ACC1, owner=Raja, balance=1500.00]

// Static members
System.out.println(BankAccount.accountCount);
```

## Records (Java 16+) — Immutable Data Classes

```java
public record Point(double x, double y) {
    // Compact constructor (validation)
    public Point {
        if (Double.isNaN(x) || Double.isNaN(y))
            throw new IllegalArgumentException("Coordinates cannot be NaN");
    }

    // Custom method
    public double distanceTo(Point other) {
        return Math.hypot(other.x - x, other.y - y);
    }
}

var p = new Point(3, 4);
p.x()              // 3.0
p.distanceTo(new Point(0, 0))  // 5.0
```

## Lab Exercise
1. Build a `Student` class with name, id, grades[] — compute average, min, max
2. Create a `Point` record and implement `distance()`, `midpoint()`, `translate()`
3. Override `equals()` and `hashCode()` and verify two equal objects in a `HashSet`
'''),

'_08_02_02_encapsulation_and_access.md': ('08_02_02','Encapsulation and Access Control','Java',2,'Object-Oriented Programming',2,'intermediate',['private','protected','public','package-private','encapsulation','getter','setter','immutable','final','record'],'''
## Access Modifiers

| Modifier | Same Class | Same Package | Subclass | Any |
|---|---|---|---|---|
| `private` | ✅ | ❌ | ❌ | ❌ |
| (package) | ✅ | ✅ | ❌ | ❌ |
| `protected` | ✅ | ✅ | ✅ | ❌ |
| `public` | ✅ | ✅ | ✅ | ✅ |

## Encapsulation Pattern

```java
public class Temperature {
    private double celsius;

    public Temperature(double celsius) {
        setCelsius(celsius);
    }

    public double getCelsius()    { return celsius; }
    public double getFahrenheit() { return celsius * 9/5 + 32; }
    public double getKelvin()     { return celsius + 273.15; }

    public void setCelsius(double celsius) {
        if (celsius < -273.15)
            throw new IllegalArgumentException("Below absolute zero!");
        this.celsius = celsius;
    }
}
```

## Immutable Classes

```java
// All fields final, no setters, defensive copies
public final class Money {
    private final double amount;
    private final String currency;

    public Money(double amount, String currency) {
        if (amount < 0) throw new IllegalArgumentException();
        this.amount = amount;
        this.currency = currency;
    }

    public double getAmount()   { return amount; }
    public String getCurrency() { return currency; }

    public Money add(Money other) {
        if (!currency.equals(other.currency)) throw new IllegalStateException();
        return new Money(amount + other.amount, currency);  // new object
    }
}
```

## Builder Pattern

```java
public class User {
    private final String name;
    private final String email;
    private final int age;

    private User(Builder b) {
        this.name  = b.name;
        this.email = b.email;
        this.age   = b.age;
    }

    public static class Builder {
        private String name;
        private String email;
        private int age = 0;

        public Builder name(String name)   { this.name = name; return this; }
        public Builder email(String email) { this.email = email; return this; }
        public Builder age(int age)        { this.age = age; return this; }
        public User build()                { return new User(this); }
    }
}

User user = new User.Builder()
    .name("Raja")
    .email("raja@example.com")
    .age(28)
    .build();
```

## Lab Exercise
1. Build an immutable `ImmutableList<T>` wrapper that throws on modification attempts
2. Implement `Address` using Builder pattern with required and optional fields
3. Show how `final` prevents subclassing and why `String` is final
'''),

'_08_02_03_inheritance.md': ('08_02_03','Inheritance','Java',2,'Object-Oriented Programming',3,'intermediate',['extends','super','override','final','abstract','Object','instanceof','covariant-return'],'''
## Inheritance Basics

```java
// Base class
public class Vehicle {
    protected String make;
    protected int year;

    public Vehicle(String make, int year) {
        this.make = make;
        this.year = year;
    }

    public String getInfo() {
        return String.format("%d %s", year, make);
    }

    public void honk() { System.out.println("Beep!"); }
}

// Subclass
public class Car extends Vehicle {
    private int doors;

    public Car(String make, int year, int doors) {
        super(make, year);   // must be first statement
        this.doors = doors;
    }

    @Override
    public String getInfo() {
        return super.getInfo() + " (" + doors + " doors)";
    }

    @Override
    public void honk() { System.out.println("Honk!"); }
}

// Usage
Vehicle v = new Car("Toyota", 2024, 4);  // polymorphism
System.out.println(v.getInfo());          // "2024 Toyota (4 doors)"
v.honk();                                 // "Honk!"
```

## Abstract Classes

```java
public abstract class Shape {
    protected String color;

    public Shape(String color) { this.color = color; }

    // Abstract — must be implemented by subclasses
    public abstract double area();
    public abstract double perimeter();

    // Concrete method (shared implementation)
    public void describe() {
        System.out.printf("%s %s: area=%.2f%n",
            color, getClass().getSimpleName(), area());
    }
}

public class Circle extends Shape {
    private double radius;

    public Circle(String color, double radius) {
        super(color);
        this.radius = radius;
    }

    @Override public double area()      { return Math.PI * radius * radius; }
    @Override public double perimeter() { return 2 * Math.PI * radius; }
}
```

## final Keyword

```java
// final class — cannot be subclassed (e.g., String, Integer)
public final class ImmutableConfig { ... }

// final method — cannot be overridden
public final void validateInput() { ... }

// final field — cannot be reassigned
private final String id = UUID.randomUUID().toString();
```

## Lab Exercise
1. Build Animal → Mammal → Dog/Cat hierarchy with abstract `makeSound()`
2. Override `toString()` at each level and verify the chain with `super.toString()`
3. Use `instanceof` with pattern matching to handle different Vehicle subtypes
'''),

'_08_02_04_polymorphism_and_abstraction.md': ('08_02_04','Polymorphism and Abstraction','Java',2,'Object-Oriented Programming',4,'intermediate',['polymorphism','dynamic-dispatch','method-overriding','abstract','sealed','interface','duck-typing'],'''
## Runtime Polymorphism

```java
// Method called depends on ACTUAL type, not declared type
Shape[] shapes = {
    new Circle("red", 5),
    new Rectangle("blue", 4, 6),
    new Triangle("green", 3, 4, 5)
};

double totalArea = 0;
for (Shape s : shapes) {
    totalArea += s.area();   // dynamic dispatch — correct area() called
    s.describe();
}
System.out.printf("Total: %.2f%n", totalArea);
```

## Sealed Classes (Java 17+)

```java
public sealed class Result<T>
    permits Result.Success, Result.Failure {

    public static final class Success<T> extends Result<T> {
        public final T value;
        public Success(T value) { this.value = value; }
    }

    public static final class Failure<T> extends Result<T> {
        public final String error;
        public Failure(String error) { this.error = error; }
    }
}

// Pattern matching switch
String message = switch (result) {
    case Result.Success<String> s -> "Got: " + s.value;
    case Result.Failure<String> f -> "Error: " + f.error;
};
```

## Abstract vs Interface

| | Abstract Class | Interface |
|---|---|---|
| Instantiate | No | No |
| Multiple inheritance | No | Yes |
| Constructor | Yes | No |
| Fields | Any | `public static final` only |
| Methods | Abstract + concrete | Default + abstract + static |
| Use when | Related classes share base | Unrelated classes share behaviour |

## Lab Exercise
1. Implement a `PaymentProcessor` hierarchy with `CreditCard`, `UPI`, `Wallet`
2. Use sealed `Result<T>` instead of checked exceptions in a file reader
3. Demonstrate how adding a new Shape subclass requires zero changes in the loop
'''),

'_08_02_05_interfaces_and_design_patterns.md': ('08_02_05','Interfaces and Design Patterns','Java',2,'Object-Oriented Programming',5,'intermediate',['interface','default-method','static-method','functional-interface','Comparable','Comparator','Singleton','Factory','Strategy'],'''
## Interfaces

```java
public interface Drawable {
    void draw();                              // abstract
    default void drawWithBorder() {           // default (Java 8+)
        System.out.println("[border]");
        draw();
    }
    static Drawable circle(double r) {        // static factory
        return () -> System.out.printf("Circle r=%.1f%n", r);
    }
}

// Functional interface (single abstract method)
@FunctionalInterface
public interface Transformer<T> {
    T transform(T input);
}

Transformer<String> upper = s -> s.toUpperCase();
upper.transform("hello");   // "HELLO"
```

## Comparable and Comparator

```java
public class Product implements Comparable<Product> {
    private String name;
    private double price;

    @Override
    public int compareTo(Product other) {
        return Double.compare(this.price, other.price);
    }
}

// Sort by price ascending
Collections.sort(products);

// Sort by name (external comparator)
products.sort(Comparator.comparing(Product::getName));

// Multi-level sort
products.sort(
    Comparator.comparing(Product::getCategory)
              .thenComparing(Product::getPrice)
              .thenComparing(Comparator.comparing(Product::getName).reversed())
);
```

## Design Patterns

```java
// Singleton
public class DatabasePool {
    private static volatile DatabasePool instance;
    private DatabasePool() {}
    public static DatabasePool getInstance() {
        if (instance == null) {
            synchronized (DatabasePool.class) {
                if (instance == null) instance = new DatabasePool();
            }
        }
        return instance;
    }
}

// Strategy
public interface SortStrategy { void sort(int[] arr); }
public class BubbleSort implements SortStrategy {
    public void sort(int[] arr) { /* ... */ }
}
public class QuickSort implements SortStrategy {
    public void sort(int[] arr) { /* ... */ }
}
public class Sorter {
    private SortStrategy strategy;
    public Sorter(SortStrategy s) { this.strategy = s; }
    public void sort(int[] arr) { strategy.sort(arr); }
}
```

## Lab Exercise
1. Define a `Logger` interface with `log(String)`, `warn(String)`, `error(String)` — implement Console and File versions
2. Sort a list of employees by department then salary using chained `Comparator`
3. Implement Strategy pattern for discount calculation: flat, percentage, buy-2-get-1
'''),

'_08_03_01_collections_framework.md': ('08_03_01','Collections Framework','Java',3,'Collections and Generics',1,'intermediate',['Collection','List','ArrayList','LinkedList','Set','HashSet','TreeSet','Map','HashMap','TreeMap','LinkedHashMap'],'''
## Collection Hierarchy

```
Collection
├── List      — ordered, duplicates allowed
│   ├── ArrayList
│   ├── LinkedList
│   └── Vector (legacy)
├── Set       — unique elements
│   ├── HashSet (unordered, O(1))
│   ├── LinkedHashSet (insertion order)
│   └── TreeSet (sorted, O(log n))
└── Queue
    ├── LinkedList
    ├── PriorityQueue
    └── ArrayDeque

Map (not Collection)
├── HashMap (unordered, O(1))
├── LinkedHashMap (insertion order)
├── TreeMap (sorted, O(log n))
└── Hashtable (legacy, synchronised)
```

## List — ArrayList vs LinkedList

```java
// ArrayList — fast random access, slow insert/delete in middle
List<String> list = new ArrayList<>();
list.add("Alice");
list.add(0, "Bob");        // insert at index
list.remove("Alice");
list.remove(0);            // remove by index
list.get(0);               // O(1)
list.size();

// LinkedList — fast insert/delete, slow random access
LinkedList<Integer> ll = new LinkedList<>();
ll.addFirst(1);   ll.addLast(3);   ll.add(1, 2);
ll.peekFirst();   ll.pollLast();

// Factory methods (immutable)
List<String> names = List.of("Alice", "Bob", "Charlie");
Set<Integer>  nums = Set.of(1, 2, 3);
Map<String,Integer> ages = Map.of("Alice", 25, "Bob", 30);
```

## Map Operations

```java
Map<String, Integer> scores = new HashMap<>();
scores.put("Alice", 95);
scores.put("Bob", 87);
scores.putIfAbsent("Carol", 0);          // only if not present
scores.merge("Alice", 5, Integer::sum);   // 95 + 5 = 100

// Get with default
int score = scores.getOrDefault("Dave", 0);

// Compute
scores.compute("Bob", (k, v) -> v == null ? 1 : v + 1);

// Iterate
for (Map.Entry<String, Integer> e : scores.entrySet()) {
    System.out.println(e.getKey() + ": " + e.getValue());
}
scores.forEach((k, v) -> System.out.printf("%s=%d%n", k, v));
```

## Lab Exercise
1. Count word frequency from a text using `HashMap` then sort by frequency using `TreeMap`
2. Demonstrate why `equals()`/`hashCode()` must be correct for `HashSet` to work
3. Implement a LRU Cache using `LinkedHashMap` with `removeEldestEntry()`
'''),

'_08_03_02_iterators_and_comparators.md': ('08_03_02','Iterators and Comparators','Java',3,'Collections and Generics',2,'intermediate',['Iterator','ListIterator','Iterable','Comparator','Comparable','Collections','sort','min','max','frequency'],'''
## Iterator Pattern

```java
List<String> items = new ArrayList<>(List.of("a","b","c","d"));

// External iterator
Iterator<String> it = items.iterator();
while (it.hasNext()) {
    String item = it.next();
    if (item.equals("b")) it.remove();  // safe removal during iteration
}

// ListIterator (bidirectional)
ListIterator<String> lit = items.listIterator();
while (lit.hasNext()) {
    String item = lit.next();
    lit.set(item.toUpperCase());  // replace current element
}

// ConcurrentModificationException — WRONG
for (String s : items) {
    if (s.equals("a")) items.remove(s);  // throws!
}
```

## Implementing Iterable

```java
public class NumberRange implements Iterable<Integer> {
    private final int start, end;
    public NumberRange(int start, int end) {
        this.start = start; this.end = end;
    }

    @Override
    public Iterator<Integer> iterator() {
        return new Iterator<>() {
            int current = start;
            public boolean hasNext() { return current <= end; }
            public Integer next()    { return current++; }
        };
    }
}

for (int n : new NumberRange(1, 5)) {
    System.out.print(n + " ");   // 1 2 3 4 5
}
```

## Collections Utility Class

```java
List<Integer> nums = new ArrayList<>(List.of(3,1,4,1,5,9,2,6));

Collections.sort(nums);
Collections.sort(nums, Comparator.reverseOrder());
Collections.shuffle(nums);
Collections.reverse(nums);
Collections.min(nums);
Collections.max(nums);
Collections.frequency(nums, 1);         // count of 1s
Collections.nCopies(3, "x");           // ["x","x","x"]
Collections.unmodifiableList(nums);     // read-only view
Collections.synchronizedList(nums);     // thread-safe wrapper
```

## Lab Exercise
1. Build a custom `CircularIterator<T>` that wraps around at the end
2. Sort a `List<Employee>` by multiple criteria: department → salary desc → name asc
3. Use `Collections.rotate(list, n)` to implement a round-robin scheduler
'''),

'_08_03_03_generics.md': ('08_03_03','Generics','Java',3,'Collections and Generics',3,'intermediate',['generics','type-parameter','wildcard','bounded','extends','super','type-erasure','generic-method'],'''
## Generic Classes

```java
public class Pair<A, B> {
    private final A first;
    private final B second;

    public Pair(A first, B second) {
        this.first = first;
        this.second = second;
    }

    public A getFirst()  { return first; }
    public B getSecond() { return second; }

    @Override public String toString() {
        return "(" + first + ", " + second + ")";
    }
}

Pair<String, Integer> p = new Pair<>("Raja", 28);
String name = p.getFirst();
```

## Generic Methods

```java
public static <T extends Comparable<T>> T max(T a, T b) {
    return a.compareTo(b) >= 0 ? a : b;
}

max(3, 7)            // 7
max("apple","mango") // "mango"
max(3.14, 2.71)      // 3.14
```

## Wildcards

```java
// Unbounded — any type
public static void printList(List<?> list) {
    for (Object item : list) System.out.println(item);
}

// Upper bounded — T or subtype (producer)
public static double sumList(List<? extends Number> list) {
    return list.stream().mapToDouble(Number::doubleValue).sum();
}
sumList(List.of(1, 2, 3));    // List<Integer> — OK
sumList(List.of(1.5, 2.5));   // List<Double>  — OK

// Lower bounded — T or supertype (consumer)
public static void addNumbers(List<? super Integer> list) {
    list.add(1); list.add(2); list.add(3);
}
addNumbers(new ArrayList<Number>());  // OK
addNumbers(new ArrayList<Object>());  // OK

// PECS: Producer Extends, Consumer Super
```

## Type Erasure

```java
// At compile time: List<String> and List<Integer> are different
// At runtime: both become List (type erased!)
List<String> strings = new ArrayList<>();
List<Integer> ints = new ArrayList<>();
System.out.println(strings.getClass() == ints.getClass()); // true!
```

## Lab Exercise
1. Build a generic `Stack<T>` with `push`, `pop`, `peek`, `isEmpty`
2. Write `filter(List<T> list, Predicate<T> pred)` — returns filtered list
3. Demonstrate PECS with a `copy(List<? super T> dest, List<? extends T> src)` method
'''),

'_08_04_01_exception_handling.md': ('08_04_01','Exception Handling','Java',4,'Exceptions and I/O',1,'intermediate',['try','catch','finally','throws','throw','checked','unchecked','custom-exception','multi-catch','try-with-resources'],'''
## Exception Hierarchy

```
Throwable
├── Error          (JVM errors — don't catch)
│   ├── OutOfMemoryError
│   └── StackOverflowError
└── Exception
    ├── RuntimeException  (unchecked — no declaration needed)
    │   ├── NullPointerException
    │   ├── IllegalArgumentException
    │   ├── IndexOutOfBoundsException
    │   └── ClassCastException
    └── Checked exceptions (must declare/catch)
        ├── IOException
        ├── SQLException
        └── ParseException
```

## try / catch / finally

```java
public double divide(double a, double b) {
    try {
        return a / b;
    } catch (ArithmeticException e) {
        System.err.println("Error: " + e.getMessage());
        return 0;
    } finally {
        System.out.println("Operation attempted");  // always runs
    }
}

// Multi-catch (Java 7+)
try {
    parseAndSave(data);
} catch (NumberFormatException | NullPointerException e) {
    throw new IllegalArgumentException("Invalid data: " + e.getMessage(), e);
}
```

## try-with-resources

```java
// Auto-closes anything implementing AutoCloseable
try (
    FileReader  fr = new FileReader("file.txt");
    BufferedReader br = new BufferedReader(fr)
) {
    String line;
    while ((line = br.readLine()) != null) {
        process(line);
    }
} catch (IOException e) {
    throw new RuntimeException("Failed to read file", e);
}
// Both br and fr are automatically closed
```

## Custom Exceptions

```java
// Checked exception
public class InsufficientFundsException extends Exception {
    private final double amount;
    public InsufficientFundsException(double amount) {
        super("Insufficient funds. Needed: " + amount);
        this.amount = amount;
    }
    public double getAmount() { return amount; }
}

// Unchecked exception
public class DuplicateUserException extends RuntimeException {
    public DuplicateUserException(String username) {
        super("User already exists: " + username);
    }
}

// Declare checked exceptions
public void withdraw(double amount) throws InsufficientFundsException {
    if (amount > balance) throw new InsufficientFundsException(amount);
    balance -= amount;
}
```

## Lab Exercise
1. Write a file parser that throws custom `ParseException` with line number and message
2. Implement `retry(Callable<T>, int times)` that retries on IOException
3. Compare checked vs unchecked — when to use each in a REST API context
'''),

'_08_04_02_file_io_and_nio.md': ('08_04_02','File I/O and NIO','Java',4,'Exceptions and I/O',2,'intermediate',['File','Path','Paths','Files','BufferedReader','BufferedWriter','FileInputStream','NIO','StandardOpenOption','walk','glob'],'''
## Classic I/O

```java
import java.io.*;

// Read text file
try (BufferedReader br = new BufferedReader(new FileReader("data.txt"))) {
    String line;
    while ((line = br.readLine()) != null) {
        System.out.println(line);
    }
}

// Write text file
try (BufferedWriter bw = new BufferedWriter(new FileWriter("output.txt"))) {
    bw.write("Line 1");
    bw.newLine();
    bw.write("Line 2");
}

// Binary I/O
try (FileInputStream fis = new FileInputStream("image.png");
     FileOutputStream fos = new FileOutputStream("copy.png")) {
    byte[] buffer = new byte[8192];
    int bytesRead;
    while ((bytesRead = fis.read(buffer)) != -1) {
        fos.write(buffer, 0, bytesRead);
    }
}
```

## NIO.2 (java.nio.file) — Modern API

```java
import java.nio.file.*;
import java.nio.charset.StandardCharsets;

Path path = Path.of("data", "file.txt");   // or Paths.get("data/file.txt")

// One-liner read
String content  = Files.readString(path, StandardCharsets.UTF_8);
List<String> lines = Files.readAllLines(path);
byte[] bytes   = Files.readAllBytes(path);

// One-liner write
Files.writeString(path, "Hello", StandardCharsets.UTF_8, StandardOpenOption.CREATE);
Files.write(path, bytes);

// Append
Files.writeString(path, "new line\n", StandardOpenOption.APPEND);

// Copy, move, delete
Files.copy(src, dst, StandardCopyOption.REPLACE_EXISTING);
Files.move(src, dst);
Files.delete(path);
Files.deleteIfExists(path);

// Directories
Files.createDirectories(Path.of("a/b/c"));
Files.exists(path)
Files.isDirectory(path)
Files.isRegularFile(path)
Files.size(path)
Files.getLastModifiedTime(path)
```

## Walking the File Tree

```java
// List all .java files recursively
try (Stream<Path> stream = Files.walk(Path.of("src"))) {
    stream.filter(p -> p.toString().endsWith(".java"))
          .forEach(System.out::println);
}

// Glob pattern matching
try (DirectoryStream<Path> ds = Files.newDirectoryStream(Path.of("."), "*.txt")) {
    for (Path p : ds) System.out.println(p);
}
```

## Lab Exercise
1. Build a file search tool: walk directory tree, find files matching a pattern, output sizes
2. Copy a directory tree recursively using `Files.walk`
3. Write a `WordCounter` that reads a file and counts unique words using streams
'''),

'_08_04_03_serialization.md': ('08_04_03','Serialization','Java',4,'Exceptions and I/O',3,'intermediate',['Serializable','ObjectOutputStream','ObjectInputStream','transient','serialVersionUID','JSON','Gson','Jackson'],'''
## Java Object Serialization

```java
import java.io.*;

// Must implement Serializable
public class User implements Serializable {
    private static final long serialVersionUID = 1L;  // version control
    private String name;
    private String email;
    private transient String password;  // NOT serialized

    // constructor, getters...
}

// Serialize (write)
User user = new User("Raja", "raja@example.com");
try (ObjectOutputStream oos = new ObjectOutputStream(
        new FileOutputStream("user.ser"))) {
    oos.writeObject(user);
}

// Deserialize (read)
try (ObjectInputStream ois = new ObjectInputStream(
        new FileInputStream("user.ser"))) {
    User loaded = (User) ois.readObject();
    System.out.println(loaded.getName());
}
```

## JSON with Gson

```java
// pom.xml: com.google.code.gson:gson:2.10.1
import com.google.gson.*;

Gson gson = new GsonBuilder()
    .setPrettyPrinting()
    .setDateFormat("yyyy-MM-dd")
    .create();

// Object → JSON string
String json = gson.toJson(user);

// JSON string → Object
User parsed = gson.fromJson(json, User.class);

// List of objects
Type listType = new TypeToken<List<User>>(){}.getType();
List<User> users = gson.fromJson(jsonArray, listType);
```

## JSON with Jackson

```java
// pom.xml: com.fasterxml.jackson.core:jackson-databind:2.17
import com.fasterxml.jackson.databind.*;

ObjectMapper mapper = new ObjectMapper();
mapper.configure(SerializationFeature.INDENT_OUTPUT, true);

// Write
String json = mapper.writeValueAsString(user);
mapper.writeValue(new File("user.json"), user);

// Read
User user = mapper.readValue(json, User.class);
User user = mapper.readValue(new File("user.json"), User.class);

// Tree model (dynamic)
JsonNode root = mapper.readTree(json);
String name = root.get("name").asText();
```

## Lab Exercise
1. Serialize a `Product` list to binary with Java serialization, then JSON with Jackson
2. Show what happens when `serialVersionUID` changes — demonstrate `InvalidClassException`
3. Build a JSON config reader that loads app settings from `config.json`
'''),

'_08_05_01_lambda_and_streams.md': ('08_05_01','Lambda Expressions and Streams','Java',5,'Modern Java',1,'intermediate',['lambda','Stream','filter','map','reduce','collect','Optional','method-reference','Collectors','flatMap'],'''
## Lambda Expressions

```java
// (params) -> expression  OR  (params) -> { body }
Runnable r = () -> System.out.println("Hello!");
Comparator<String> byLength = (a, b) -> Integer.compare(a.length(), b.length());
Predicate<Integer> isEven = n -> n % 2 == 0;
Function<String, Integer> strLen = String::length;  // method reference

// Common functional interfaces
Predicate<T>   test(T) → boolean
Function<T,R>  apply(T) → R
Consumer<T>    accept(T) → void
Supplier<T>    get() → T
BiFunction<T,U,R> apply(T,U) → R
UnaryOperator<T>  apply(T) → T
BinaryOperator<T> apply(T,T) → T
```

## Stream API

```java
import java.util.stream.*;

List<String> names = List.of("Alice","Bob","Charlie","Dave","Eve");

// filter, map, collect
List<String> result = names.stream()
    .filter(n -> n.length() > 3)      // filter: Charlie, Dave
    .map(String::toUpperCase)          // transform
    .sorted()                          // sort alphabetically
    .collect(Collectors.toList());     // collect to list

// Numeric operations
IntStream.range(1, 11).sum()       // 55
IntStream.of(1,2,3).average()      // OptionalDouble[2.0]

List<Integer> nums = List.of(1,2,3,4,5,6,7,8,9,10);
int sumOfSquaresOfEvens = nums.stream()
    .filter(n -> n % 2 == 0)
    .mapToInt(n -> n * n)
    .sum();   // 220
```

## Collectors

```java
// Grouping
Map<Integer, List<String>> byLength = names.stream()
    .collect(Collectors.groupingBy(String::length));

// Counting per group
Map<Integer, Long> counts = names.stream()
    .collect(Collectors.groupingBy(String::length, Collectors.counting()));

// Joining
String joined = names.stream()
    .collect(Collectors.joining(", ", "[", "]"));  // "[Alice, Bob, ...]"

// Partitioning
Map<Boolean, List<Integer>> partitioned = nums.stream()
    .collect(Collectors.partitioningBy(n -> n % 2 == 0));
```

## Optional

```java
Optional<String> opt = Optional.of("Hello");
Optional<String> empty = Optional.empty();

opt.isPresent()           // true
opt.get()                 // "Hello" (throws if empty)
opt.orElse("default")     // returns value or "default"
opt.orElseGet(() -> compute_default())
opt.map(String::toUpperCase)  // Optional["HELLO"]
opt.filter(s -> s.length() > 3)
opt.ifPresent(System.out::println)

// Safe chaining
Optional<String> result = Optional.ofNullable(user)
    .map(User::getAddress)
    .map(Address::getCity)
    .filter(city -> city.startsWith("C"));
```

## Lab Exercise
1. Count word frequencies in a paragraph using streams and `Collectors.groupingBy`
2. Find top 3 most expensive products per category using streams
3. Flatten a `List<List<Integer>>` to `List<Integer>` using `flatMap`
'''),

'_08_05_02_java_8_to_21_features.md': ('08_05_02','Java 8 to 21 Key Features','Java',5,'Modern Java',2,'intermediate',['records','sealed','text-blocks','switch-expression','pattern-matching','var','instanceof','virtual-threads'],'''
## Java Version Feature Map

| Version | Key Feature |
|---|---|
| 8 | Lambda, Stream, Optional, Date/Time API |
| 10 | `var` local type inference |
| 11 | `String.lines()`, `isBlank()`, `strip()` |
| 14 | Records (preview), Switch expressions |
| 15 | Text blocks |
| 16 | Records (final), `instanceof` pattern matching |
| 17 | Sealed classes, LTS |
| 21 | Virtual threads (Loom), Pattern matching switch, LTS |

## Records (Java 16)

```java
record Person(String name, int age) {
    // Compact constructor — validation
    Person {
        Objects.requireNonNull(name);
        if (age < 0) throw new IllegalArgumentException();
    }

    // Custom methods
    String greeting() { return "Hello, " + name + "!"; }
}

var p = new Person("Raja", 28);
p.name()      // "Raja" (auto-generated accessor)
p.age()       // 28
p.equals(new Person("Raja", 28))   // true (auto-generated)
```

## Sealed Classes (Java 17)

```java
sealed interface Shape permits Circle, Rectangle, Triangle {}

record Circle(double radius) implements Shape {}
record Rectangle(double w, double h) implements Shape {}
record Triangle(double a, double b, double c) implements Shape {}

double area(Shape s) {
    return switch (s) {
        case Circle c    -> Math.PI * c.radius() * c.radius();
        case Rectangle r -> r.w() * r.h();
        case Triangle t  -> {
            double sp = (t.a()+t.b()+t.c()) / 2;
            yield Math.sqrt(sp*(sp-t.a())*(sp-t.b())*(sp-t.c()));
        }
    };
}
```

## Virtual Threads (Java 21)

```java
// Classic: 1 OS thread per request — limited scalability
ExecutorService exec = Executors.newFixedThreadPool(200);

// Virtual threads: millions of lightweight threads
ExecutorService vExec = Executors.newVirtualThreadPerTaskExecutor();

// Per-request virtual thread
try (var executor = Executors.newVirtualThreadPerTaskExecutor()) {
    IntStream.range(0, 10_000).forEach(i ->
        executor.submit(() -> {
            Thread.sleep(Duration.ofMillis(100));  // non-blocking
            return i;
        })
    );
}
```

## Lab Exercise
1. Rewrite a `User` class as a record, add validation in compact constructor
2. Implement a `Shape` sealed interface with pattern-matching area computation
3. Compare throughput: 100 requests with thread pool vs virtual threads
'''),

'_08_05_03_concurrency_and_threading.md': ('08_05_03','Concurrency and Threading','Java',5,'Modern Java',3,'advanced',['Thread','Runnable','ExecutorService','Future','CompletableFuture','synchronized','ReentrantLock','volatile','ConcurrentHashMap'],'''
## Thread Basics

```java
// Extend Thread
class PrintTask extends Thread {
    @Override
    public void run() { System.out.println("Running: " + getName()); }
}
new PrintTask().start();

// Implement Runnable (preferred)
Runnable task = () -> System.out.println("Task: " + Thread.currentThread().getName());
Thread t = new Thread(task, "my-thread");
t.start();
t.join();   // wait for completion
```

## ExecutorService

```java
ExecutorService exec = Executors.newFixedThreadPool(4);

// Submit tasks
Future<Integer> future = exec.submit(() -> expensiveCompute());
Integer result = future.get(5, TimeUnit.SECONDS);  // blocking get with timeout

// ScheduledExecutorService
ScheduledExecutorService scheduler = Executors.newScheduledThreadPool(1);
scheduler.scheduleAtFixedRate(() -> pollDatabase(), 0, 30, TimeUnit.SECONDS);

exec.shutdown();        // no more tasks, let existing finish
exec.awaitTermination(60, TimeUnit.SECONDS);
exec.shutdownNow();     // interrupt running tasks
```

## CompletableFuture

```java
CompletableFuture<String> future = CompletableFuture
    .supplyAsync(() -> fetchData("https://api.example.com"))
    .thenApply(data -> parseJson(data))
    .thenApply(obj -> obj.getName())
    .exceptionally(ex -> "default-name");

// Combine multiple
CompletableFuture<String> f1 = fetchAsync("url1");
CompletableFuture<String> f2 = fetchAsync("url2");

CompletableFuture.allOf(f1, f2).thenRun(() -> {
    System.out.println(f1.join() + f2.join());
});
```

## Synchronization

```java
// synchronized method
public synchronized void increment() { count++; }

// synchronized block
public void process() {
    synchronized (this) { count++; }
}

// ReentrantLock
private final ReentrantLock lock = new ReentrantLock();
public void safeIncrement() {
    lock.lock();
    try { count++; }
    finally { lock.unlock(); }
}

// Atomic variables (lock-free)
private final AtomicInteger count = new AtomicInteger(0);
count.incrementAndGet();

// ConcurrentHashMap
ConcurrentHashMap<String, Integer> map = new ConcurrentHashMap<>();
map.merge("key", 1, Integer::sum);
```

## Lab Exercise
1. Build a thread-safe counter with `AtomicInteger` and verify with 100 concurrent threads
2. Fetch 10 URLs concurrently using `CompletableFuture.allOf`, combine results
3. Implement a bounded blocking queue using `ReentrantLock` + `Condition`
'''),

'_08_06_01_jdbc_fundamentals.md': ('08_06_01','JDBC Fundamentals','Java',6,'Database Access',1,'intermediate',['JDBC','Connection','Statement','PreparedStatement','ResultSet','DataSource','connection-pool','HikariCP','transaction'],'''
## JDBC Basics

```java
import java.sql.*;

// Load driver (auto-registered in JDBC 4+)
String url = "jdbc:mysql://localhost:3306/mydb";
String user = "root", pass = "password";

try (Connection conn = DriverManager.getConnection(url, user, pass)) {
    // PreparedStatement — prevents SQL injection
    String sql = "SELECT id, name, salary FROM employees WHERE dept_id = ?";
    try (PreparedStatement ps = conn.prepareStatement(sql)) {
        ps.setInt(1, 3);
        try (ResultSet rs = ps.executeQuery()) {
            while (rs.next()) {
                int id       = rs.getInt("id");
                String name  = rs.getString("name");
                double salary = rs.getDouble("salary");
                System.out.printf("%d: %s = %.2f%n", id, name, salary);
            }
        }
    }
}
```

## CRUD Operations

```java
// INSERT with generated key
String insert = "INSERT INTO products (name, price) VALUES (?, ?)";
try (PreparedStatement ps = conn.prepareStatement(insert,
        Statement.RETURN_GENERATED_KEYS)) {
    ps.setString(1, "Widget");
    ps.setDouble(2, 9.99);
    int rows = ps.executeUpdate();   // returns affected rows
    try (ResultSet keys = ps.getGeneratedKeys()) {
        if (keys.next()) System.out.println("New ID: " + keys.getInt(1));
    }
}

// UPDATE / DELETE
String update = "UPDATE products SET price = ? WHERE id = ?";
try (PreparedStatement ps = conn.prepareStatement(update)) {
    ps.setDouble(1, 14.99);
    ps.setInt(2, 5);
    ps.executeUpdate();
}
```

## Connection Pooling with HikariCP

```java
import com.zaxxer.hikari.*;

HikariConfig config = new HikariConfig();
config.setJdbcUrl("jdbc:mysql://localhost:3306/mydb");
config.setUsername("root");
config.setPassword("password");
config.setMaximumPoolSize(10);
config.setMinimumIdle(5);
config.setConnectionTimeout(30000);

HikariDataSource ds = new HikariDataSource(config);

try (Connection conn = ds.getConnection()) {
    // use conn
}  // returned to pool automatically
```

## Transactions

```java
conn.setAutoCommit(false);
try {
    deductBalance(conn, fromId, amount);
    addBalance(conn, toId, amount);
    conn.commit();
} catch (SQLException e) {
    conn.rollback();
    throw e;
}
```

## Lab Exercise
1. Build a `ProductDAO` with CRUD methods using `PreparedStatement`
2. Configure HikariCP and benchmark single connection vs pool for 100 queries
3. Implement a bank transfer with proper transaction rollback on any failure
'''),

'_08_06_02_jpa_and_hibernate.md': ('08_06_02','JPA and Hibernate','Java',6,'Database Access',2,'advanced',['JPA','Hibernate','Entity','EntityManager','Repository','JPQL','HQL','OneToMany','ManyToOne','lazy','eager'],'''
## JPA Entities

```java
import jakarta.persistence.*;

@Entity
@Table(name = "employees")
public class Employee {
    @Id
    @GeneratedValue(strategy = GenerationType.IDENTITY)
    private Long id;

    @Column(name = "first_name", nullable = false, length = 50)
    private String firstName;

    @Column(unique = true)
    private String email;

    @ManyToOne(fetch = FetchType.LAZY)
    @JoinColumn(name = "dept_id")
    private Department department;

    @OneToMany(mappedBy = "employee", cascade = CascadeType.ALL,
               orphanRemoval = true)
    private List<Skill> skills = new ArrayList<>();

    // constructors, getters, setters...
}
```

## EntityManager CRUD

```java
@PersistenceContext
EntityManager em;

// Create
em.persist(employee);

// Read
Employee e = em.find(Employee.class, 1L);

// Update
Employee e = em.find(Employee.class, 1L);
e.setSalary(75000);   // auto-tracked in active transaction

// Delete
em.remove(em.find(Employee.class, 1L));

// JPQL
List<Employee> highEarners = em.createQuery(
    "SELECT e FROM Employee e WHERE e.salary > :threshold", Employee.class)
    .setParameter("threshold", 70000.0)
    .getResultList();
```

## Spring Data JPA

```java
// Just define the interface!
public interface EmployeeRepository extends JpaRepository<Employee, Long> {
    List<Employee> findByDepartmentName(String name);
    List<Employee> findBySalaryBetween(double min, double max);

    @Query("SELECT e FROM Employee e WHERE e.salary > :min ORDER BY e.salary DESC")
    List<Employee> findHighEarners(@Param("min") double minSalary);
}

// Usage in service
@Service
@Transactional
public class EmployeeService {
    @Autowired
    private EmployeeRepository repo;

    public Employee hire(String name, String email) {
        return repo.save(new Employee(name, email));
    }

    public List<Employee> getHighEarners(double min) {
        return repo.findHighEarners(min);
    }
}
```

## Lab Exercise
1. Create a `Product`/`Category` JPA entity relationship with `@OneToMany`
2. Write a `ProductRepository` with custom JPQL queries for search and price filter
3. Compare lazy vs eager loading — show N+1 problem and fix with `@EntityGraph`
'''),
}

for fname, data in java_lessons.items():
    lid, title, course, mod, mod_title, les, diff, tags, body = data
    write(J, fname, fm(lid, title, course, mod, mod_title, les, diff, tags) + body.strip() + '\n')

print(f'Java written so far: {written}')

# ═══════════════════════════════════════════════════════════════
# C — 16 lessons
# ═══════════════════════════════════════════════════════════════
print()
print('='*60)
print('C — 16 lessons')
print('='*60)
C_DIR = '_09_c'

c_lessons = {
'_02_01_c_curriculum_placeholder.md': None,  # delete / skip placeholder
'_09_01_01_c_introduction_and_toolchain.md': ('09_01_01','C Introduction and Toolchain','C',1,'C Fundamentals',1,'beginner',['C','gcc','clang','compilation','linking','preprocessor','make','CMake','valgrind'],'''
## What is C?

C is a **procedural, statically-typed, compiled** systems programming language developed by Dennis Ritchie at Bell Labs (1972). It remains the foundation of operating systems, embedded systems, and high-performance software.

## Why Learn C?

- Deep understanding of memory management
- Foundation for C++, Java, Rust, Go
- Required for embedded/systems programming
- Close-to-hardware control

## Toolchain

```bash
# Install GCC (Linux)
sudo apt install gcc build-essential

# Compile and run
gcc -o hello hello.c
./hello

# With warnings (always use!)
gcc -Wall -Wextra -Werror -o hello hello.c

# Clang (alternative)
clang -o hello hello.c
```

## Compilation Stages

```
hello.c (source)
  ↓ Preprocessor (cpp)   → expands #include, #define
hello.i (preprocessed)
  ↓ Compiler (cc1)       → generates assembly
hello.s (assembly)
  ↓ Assembler (as)       → generates object code
hello.o (object file)
  ↓ Linker (ld)          → links libraries
hello (executable)
```

## Hello World

```c
#include <stdio.h>   /* standard I/O */
#include <stdlib.h>  /* EXIT_SUCCESS, EXIT_FAILURE */

int main(void) {
    printf("Hello, World!\\n");
    return EXIT_SUCCESS;   /* 0 */
}
```

## Makefile

```makefile
CC = gcc
CFLAGS = -Wall -Wextra -std=c11

all: myprogram

myprogram: main.o utils.o
	$(CC) $(CFLAGS) -o $@ $^

main.o: main.c main.h
	$(CC) $(CFLAGS) -c $<

clean:
	rm -f *.o myprogram
```

## Lab Exercise
1. Install GCC, compile `hello.c`, run it
2. Break compilation into stages: `-E` (preprocess), `-S` (compile), `-c` (assemble)
3. Write a `Makefile` for a two-file C project
'''),

'_09_01_02_data_types_and_operators.md': ('09_01_02','Data Types and Operators','C',1,'C Fundamentals',2,'beginner',['int','char','float','double','short','long','unsigned','sizeof','printf','scanf','operators','casting'],'''
## Primitive Types

```c
#include <stdio.h>
#include <limits.h>
#include <float.h>

/* Integer types */
char   c = 'A';         /* 1 byte  (-128 to 127) */
short  s = 30000;       /* 2 bytes */
int    i = 2147483647;  /* 4 bytes (typical) */
long   l = 9223372036854775807L;  /* 8 bytes on 64-bit */
long long ll = -9223372036854775807LL - 1;

/* Unsigned variants */
unsigned int  ui = 4294967295U;
unsigned char uc = 255;

/* Floating point */
float  f = 3.14f;
double d = 3.14159265358979;
long double ld = 3.14159265358979323846L;

/* sizeof operator */
printf("int: %zu bytes\\n", sizeof(int));    /* zu = size_t */
printf("double: %zu bytes\\n", sizeof(double));
```

## printf Format Specifiers

```c
printf("%d\\n",  42);         /* int */
printf("%ld\\n", 123456789L); /* long */
printf("%u\\n",  42u);        /* unsigned */
printf("%f\\n",  3.14f);      /* float/double (default 6 decimal) */
printf("%.2f\\n",3.14159);    /* 2 decimal places */
printf("%e\\n",  1.23e10);    /* scientific */
printf("%c\\n",  'A');        /* char */
printf("%s\\n",  "hello");    /* string */
printf("%p\\n",  &x);         /* pointer address */
printf("%x\\n",  255);        /* hex: ff */
printf("%-10s|\\n","left");   /* left-align 10 chars */
```

## scanf (Input)

```c
int age;
double salary;
char name[50];

printf("Enter age: ");
scanf("%d", &age);           /* & for address */

printf("Enter name: ");
scanf("%49s", name);         /* limit string length */

printf("Enter salary: ");
scanf("%lf", &salary);       /* %lf for double */
```

## Operators

```c
/* Arithmetic */
5 / 2     /* 2 (integer division) */
5 % 2     /* 1 */
5.0 / 2   /* 2.5 */

/* Bitwise */
0xFF & 0x0F   /* 0x0F (AND) */
0xF0 | 0x0F   /* 0xFF (OR) */
0xFF ^ 0xF0   /* 0x0F (XOR) */
~0x00         /* 0xFF...FF (NOT) */
1 << 4        /* 16  (left shift) */
256 >> 2      /* 64  (right shift) */

/* Increment / decrement */
i++;   /* post-increment: use then increment */
++i;   /* pre-increment: increment then use */

/* Cast */
int x = (int)3.99;   /* 3 — truncates */
double y = (double)5 / 2;   /* 2.5 */
```

## Lab Exercise
1. Print size of all primitive types on your system using `sizeof`
2. Read two integers, print their sum, difference, product, quotient, remainder
3. Demonstrate integer overflow: what happens when you add 1 to `INT_MAX`?
'''),

'_09_01_03_control_flow.md': ('09_01_03','Control Flow','C',1,'C Fundamentals',3,'beginner',['if','else','switch','for','while','do-while','break','continue','goto'],'''
## Conditional Statements

```c
int score = 82;

if (score >= 90) {
    printf("A\\n");
} else if (score >= 75) {
    printf("B\\n");
} else if (score >= 60) {
    printf("C\\n");
} else {
    printf("F\\n");
}

/* Ternary */
const char *result = (score >= 60) ? "Pass" : "Fail";

/* Switch */
char grade = 'B';
switch (grade) {
    case 'A': printf("Excellent\\n"); break;
    case 'B': printf("Good\\n");      break;
    case 'C': printf("Average\\n");   break;
    default:  printf("Below\\n");     break;
}
```

## Loops

```c
/* for loop */
for (int i = 0; i < 10; i++) {
    printf("%d ", i);
}

/* while loop */
int n = 1;
while (n <= 1000) {
    n *= 2;
}

/* do-while (runs at least once) */
int input;
do {
    printf("Enter positive number: ");
    scanf("%d", &input);
} while (input <= 0);

/* Nested loops — multiplication table */
for (int i = 1; i <= 9; i++) {
    for (int j = 1; j <= 9; j++) {
        printf("%3d", i * j);
    }
    printf("\\n");
}
```

## Break, Continue, Goto

```c
/* break — exit current loop */
for (int i = 0; i < 100; i++) {
    if (i * i > 1000) break;
    printf("%d\\n", i);
}

/* continue — skip to next iteration */
for (int i = 0; i < 20; i++) {
    if (i % 2 == 0) continue;
    printf("%d ", i);  /* odd numbers only */
}

/* goto — use sparingly (error cleanup is a valid use) */
int *p = malloc(100 * sizeof(int));
if (!p) goto cleanup;

int *q = malloc(200 * sizeof(int));
if (!q) goto cleanup;

/* ... use p and q ... */

cleanup:
    free(p);
    free(q);
```

## Lab Exercise
1. Print all prime numbers from 2 to 100 using nested loops and `break`
2. Implement a simple calculator with `switch` for +, -, *, /
3. Use `goto` for multi-resource cleanup in an error path
'''),

'_09_01_04_functions_and_scope.md': ('09_01_04','Functions and Scope','C',1,'C Fundamentals',4,'beginner',['function','prototype','return','void','scope','static','extern','auto','register','recursion'],'''
## Functions

```c
#include <stdio.h>

/* Function prototype (declaration) — before main */
double calculate_bmi(double weight, double height);
void print_bmi(double bmi);

int main(void) {
    double bmi = calculate_bmi(70.0, 1.75);
    print_bmi(bmi);
    return 0;
}

/* Function definitions */
double calculate_bmi(double weight, double height) {
    if (height <= 0.0) return -1.0;  /* error sentinel */
    return weight / (height * height);
}

void print_bmi(double bmi) {
    const char *category;
    if      (bmi < 18.5) category = "Underweight";
    else if (bmi < 25.0) category = "Normal";
    else if (bmi < 30.0) category = "Overweight";
    else                  category = "Obese";
    printf("BMI: %.2f (%s)\\n", bmi, category);
}
```

## Scope and Storage Classes

```c
int global_var = 10;   /* file scope, static storage */

void func(void) {
    int local = 20;    /* block scope, automatic storage */
    static int count = 0;  /* block scope, STATIC storage (persists) */
    count++;
    printf("Called %d times\\n", count);
}

/* static at file level — limits visibility to this file */
static void internal_helper(void) { /* not visible outside */ }

/* extern — access global from another file */
extern int shared_counter;
```

## Recursion

```c
/* Factorial */
unsigned long long factorial(int n) {
    if (n <= 1) return 1;
    return n * factorial(n - 1);
}

/* Fibonacci with memoization */
#define MAX 100
long long memo[MAX] = {0};
long long fib(int n) {
    if (n <= 1) return n;
    if (memo[n]) return memo[n];
    memo[n] = fib(n-1) + fib(n-2);
    return memo[n];
}

/* Tower of Hanoi */
void hanoi(int n, char from, char to, char via) {
    if (n == 1) { printf("Move disk 1 from %c to %c\\n", from, to); return; }
    hanoi(n-1, from, via, to);
    printf("Move disk %d from %c to %c\\n", n, from, to);
    hanoi(n-1, via, to, from);
}
```

## Lab Exercise
1. Write a recursive `power(base, exp)` function without using `math.h`
2. Use `static` local variable to count function invocations
3. Split a C program into two files with a header; use `extern` for shared data
'''),

'_09_02_01_arrays_and_multidimensional.md': ('09_02_01','Arrays and Multidimensional Arrays','C',2,'Memory and Data Structures',1,'beginner',['array','index','sizeof','2D-array','VLA','array-decay','function-parameter'],'''
## Arrays in C

```c
#include <stdio.h>
#include <string.h>

/* Declaration and initialisation */
int nums[5] = {10, 20, 30, 40, 50};
int zeros[100] = {0};       /* all elements initialised to 0 */
int auto_size[] = {1,2,3};  /* compiler counts: size = 3 */

int len = sizeof(nums) / sizeof(nums[0]);   /* 5 */

/* Access (0-indexed) */
printf("%d\\n", nums[0]);    /* 10 */
printf("%d\\n", nums[4]);    /* 50 */

/* Traverse */
for (int i = 0; i < len; i++) {
    printf("%d ", nums[i]);
}
```

## 2D Arrays

```c
int matrix[3][4] = {
    {1,  2,  3,  4},
    {5,  6,  7,  8},
    {9, 10, 11, 12}
};

printf("%d\\n", matrix[1][2]);  /* 7 */

/* Row-major storage (row by row in memory) */
int rows = 3, cols = 4;
for (int i = 0; i < rows; i++) {
    for (int j = 0; j < cols; j++) {
        printf("%3d", matrix[i][j]);
    }
    printf("\\n");
}
```

## Array Decay and Function Parameters

```c
/* Arrays decay to pointer when passed to functions */
void print_array(int arr[], int len) {   /* same as int *arr */
    for (int i = 0; i < len; i++) {
        printf("%d ", arr[i]);
    }
}

/* CANNOT get size from pointer inside function */
/* Must pass length explicitly */
void sort(int *arr, int len) {
    /* bubblesort */
    for (int i = 0; i < len-1; i++)
        for (int j = 0; j < len-1-i; j++)
            if (arr[j] > arr[j+1]) {
                int tmp = arr[j]; arr[j] = arr[j+1]; arr[j+1] = tmp;
            }
}
```

## Lab Exercise
1. Find maximum, minimum, and average of a 10-element array
2. Implement matrix multiplication for two 3×3 matrices
3. Search an element in a sorted array using binary search
'''),

'_09_02_02_strings_in_c.md': ('09_02_02','Strings in C','C',2,'Memory and Data Structures',2,'beginner',['char-array','null-terminator','strlen','strcpy','strcat','strcmp','sprintf','sscanf','strstr'],'''
## C Strings

A C string is a **null-terminated `char` array** — every string ends with `\0`.

```c
#include <stdio.h>
#include <string.h>

char name[20] = "Raja";           /* stored as: R a j a \\0 */
char greeting[] = "Hello";         /* size = 6 (includes \\0) */
const char *literal = "World";     /* string literal (read-only) */

printf("Length: %zu\\n", strlen(name));   /* 4 (not counting \\0) */
printf("Size:   %zu\\n", sizeof(name));   /* 20 */
```

## String Functions (string.h)

```c
char src[] = "Hello";
char dst[20];

/* Copy */
strcpy(dst, src);               /* dangerous — no bound check */
strncpy(dst, src, sizeof(dst)-1);  /* safer */
dst[sizeof(dst)-1] = '\\0';         /* ensure null termination */

/* Concatenate */
strcat(dst, " World");
strncat(dst, " World", 6);

/* Compare */
strcmp("abc", "abd")    /* negative (a<b alphabetically) */
strncmp("abc","abd", 2) /* 0 (first 2 chars equal) */

/* Search */
char *pos = strstr(name, "aj");  /* pointer to "aja" in "Raja" */
char *chr = strchr(name, 'j');   /* pointer to 'j' */

/* Format to string */
char buf[100];
snprintf(buf, sizeof(buf), "Name: %s, Age: %d", name, 28);
```

## Reading Strings

```c
char line[256];

/* Safe way — fgets reads entire line including spaces */
fgets(line, sizeof(line), stdin);
/* Remove trailing newline */
line[strcspn(line, "\\n")] = '\\0';

/* sscanf — parse formatted string */
int day, month, year;
sscanf("2024-07-28", "%d-%d-%d", &year, &month, &day);
```

## Lab Exercise
1. Implement `my_strlen`, `my_strcpy`, `my_strcat` without using `string.h`
2. Reverse a string in-place using two pointers
3. Check if a string is a palindrome (ignoring case and spaces)
'''),

'_09_02_03_pointers_fundamentals.md': ('09_02_03','Pointers Fundamentals','C',2,'Memory and Data Structures',3,'intermediate',['pointer','address','dereference','NULL','void-pointer','pointer-arithmetic','const-pointer'],'''
## What is a Pointer?

A pointer is a variable that stores the **memory address** of another variable.

```c
#include <stdio.h>

int x = 42;
int *p = &x;      /* p holds the address of x */

printf("x     = %d\\n", x);     /* 42 */
printf("&x    = %p\\n", (void*)&x);   /* address, e.g. 0x7ffee */
printf("p     = %p\\n", (void*)p);    /* same address */
printf("*p    = %d\\n", *p);    /* 42 (dereference) */

*p = 100;          /* modify x through pointer */
printf("x now = %d\\n", x);     /* 100 */
```

## Pointer Sizes and Types

```c
sizeof(int *)    /* 8 bytes on 64-bit system (all pointers same size) */
sizeof(char *)   /* 8 bytes */
sizeof(double *) /* 8 bytes */

/* void pointer — generic pointer */
void *generic = &x;
int *back = (int *)generic;   /* must cast back before use */
```

## Null and Uninitialized Pointers

```c
int *p = NULL;    /* explicitly null — safe to check */
if (p != NULL) {
    *p = 42;      /* safe */
}

/* NEVER do this */
int *bad;         /* uninitialized — undefined behaviour! */
*bad = 42;        /* may crash or corrupt memory */
```

## const and Pointers

```c
const int x = 10;
const int *p = &x;   /* pointer to const int — can't change *p */
int * const q = &y;  /* const pointer — can't change q itself */
const int * const r = &x;  /* both const */

/* Practical: protect function parameters */
void print_string(const char *str) {
    /* str cannot be modified here */
    printf("%s\\n", str);
}
```

## Pointer Arithmetic

```c
int arr[] = {10, 20, 30, 40, 50};
int *p = arr;        /* points to arr[0] */

p + 1;               /* points to arr[1] (adds sizeof(int)) */
*(p + 2);            /* arr[2] = 30 */

p++;                 /* advance to next element */
printf("%d\\n", *p); /* 20 */

/* Difference between pointers */
int *start = arr;
int *end   = &arr[4];
ptrdiff_t count = end - start;   /* 4 */
```

## Lab Exercise
1. Write `swap(int *a, int *b)` — verify it actually swaps the caller's variables
2. Traverse an array using pointer arithmetic (no `[]` subscript)
3. Explain the difference between `const int *p` and `int * const p`
'''),

'_09_02_04_pointers_and_arrays.md': ('09_02_04','Pointers and Arrays','C',2,'Memory and Data Structures',4,'intermediate',['pointer-array-equivalence','array-decay','double-pointer','pointer-to-pointer','string-array','argv'],'''
## Array-Pointer Equivalence

```c
int arr[] = {1, 2, 3, 4, 5};
int *p = arr;    /* arr decays to &arr[0] */

arr[2]   == *(arr + 2)   /* true */
arr[i]   == *(arr + i)   /* true */
p[i]     == *(p + i)     /* true */

/* But arr itself is NOT a pointer — it's an array */
sizeof(arr)  /* 20 (5 * 4) */
sizeof(p)    /* 8 (pointer size) */
/* arr = p; */  /* ERROR: arr is not assignable */
```

## Function Pointers

```c
/* Pointer to function returning int, taking two ints */
int add(int a, int b) { return a + b; }
int sub(int a, int b) { return a - b; }

int (*op)(int, int) = add;
printf("%d\\n", op(3, 4));   /* 7 */
op = sub;
printf("%d\\n", op(3, 4));   /* -1 */

/* Array of function pointers (dispatch table) */
int (*operations[])(int,int) = {add, sub, mul, div};
printf("%d\\n", operations[2](3, 4));   /* 12 */

/* Callback pattern */
void apply(int *arr, int len, int (*transform)(int)) {
    for (int i = 0; i < len; i++)
        arr[i] = transform(arr[i]);
}
int square(int x) { return x * x; }
apply(arr, 5, square);
```

## Pointer to Pointer (double pointer)

```c
int x = 42;
int *p = &x;
int **pp = &p;

**pp         /* 42 */
*pp          /* p (address of x) */

/* Common use: modify pointer in function */
void allocate(int **ptr, int size) {
    *ptr = malloc(size * sizeof(int));
}

int *data = NULL;
allocate(&data, 100);
```

## Array of Strings

```c
/* Array of string literals */
const char *days[] = {"Mon","Tue","Wed","Thu","Fri","Sat","Sun"};
printf("%s\\n", days[0]);   /* Mon */

/* main argc/argv */
int main(int argc, char *argv[]) {
    for (int i = 0; i < argc; i++) {
        printf("arg[%d] = %s\\n", i, argv[i]);
    }
}
```

## Lab Exercise
1. Implement `qsort` using a function pointer comparator for an array of structs
2. Write `strdup` equivalent that allocates and returns a copy of a string
3. Build a command dispatch table mapping command strings to functions
'''),

'_09_03_01_dynamic_memory.md': ('09_03_01','Dynamic Memory Allocation','C',3,'Advanced C',1,'intermediate',['malloc','calloc','realloc','free','heap','valgrind','memory-leak','dangling-pointer','double-free'],'''
## Heap vs Stack

```
Stack: automatic, fast, limited size (~1-8 MB)
       local variables, function frames, freed on return

Heap:  manual, slower, large (limited by RAM)
       malloc/calloc/realloc, must free() manually
```

## Allocation Functions

```c
#include <stdlib.h>

/* malloc — allocate n bytes (uninitialized) */
int *arr = malloc(10 * sizeof(int));
if (arr == NULL) {
    fprintf(stderr, "malloc failed\\n");
    exit(EXIT_FAILURE);
}

/* calloc — allocate n*size bytes (zero-initialized) */
int *zeroed = calloc(10, sizeof(int));

/* realloc — resize existing allocation */
arr = realloc(arr, 20 * sizeof(int));
if (arr == NULL) { /* original freed on failure! store backup */ }

/* free — release memory */
free(arr);
arr = NULL;   /* prevent dangling pointer */
```

## Common Mistakes

```c
/* Memory leak — allocated but never freed */
for (int i = 0; i < 1000; i++) {
    char *s = malloc(100);
    strcpy(s, "leaked string");
    /* forgot free(s)! */
}

/* Double free — undefined behaviour */
free(p);
free(p);    /* CRASH */

/* Dangling pointer — accessing freed memory */
free(p);
printf("%d\\n", *p);  /* undefined behaviour */

/* Buffer overflow — writing past allocation */
int *arr = malloc(5 * sizeof(int));
arr[5] = 99;   /* off by one — undefined behaviour */
```

## Valgrind (Memory Debugger)

```bash
gcc -g -o program program.c
valgrind --leak-check=full --track-origins=yes ./program
```

## Dynamic Array Implementation

```c
typedef struct {
    int *data;
    int size;
    int capacity;
} DynArray;

DynArray *dynarray_create(int initial) {
    DynArray *a = malloc(sizeof(DynArray));
    a->data = malloc(initial * sizeof(int));
    a->size = 0;
    a->capacity = initial;
    return a;
}

void dynarray_push(DynArray *a, int val) {
    if (a->size == a->capacity) {
        a->capacity *= 2;
        a->data = realloc(a->data, a->capacity * sizeof(int));
    }
    a->data[a->size++] = val;
}

void dynarray_free(DynArray *a) {
    free(a->data);
    free(a);
}
```

## Lab Exercise
1. Implement a dynamic string builder with `realloc` (grow by 2x when full)
2. Find all memory leaks in a provided buggy program using Valgrind
3. Build a `stack_t` using dynamic allocation with push/pop/peek
'''),

'_09_03_02_structures_and_unions.md': ('09_03_02','Structures and Unions','C',3,'Advanced C',2,'intermediate',['struct','union','typedef','padding','bit-fields','enum','nested-struct'],'''
## Structures

```c
#include <stdio.h>

/* Define structure */
typedef struct {
    char name[50];
    int  age;
    float salary;
} Employee;

/* Declaration and initialization */
Employee emp = {"Raja", 28, 75000.0f};
Employee emp2 = {.name = "Alice", .age = 30};   /* designated initialiser */

/* Access members */
printf("%s earns %.2f\\n", emp.name, emp.salary);

/* Pointer to struct */
Employee *ptr = &emp;
printf("%s\\n", ptr->name);     /* arrow operator */
printf("%d\\n", (*ptr).age);    /* same as ptr->age */
```

## Memory Layout and Padding

```c
struct Padded {
    char  a;       /* 1 byte */
    /* 3 bytes padding */
    int   b;       /* 4 bytes */
    char  c;       /* 1 byte */
    /* 3 bytes padding */
};  /* total: 12 bytes (not 6!) */

struct Packed {
    char  a;
    char  c;
    int   b;
};  /* total: 8 bytes (better packing) */

printf("%zu\\n", sizeof(struct Padded));  /* 12 */
printf("%zu\\n", sizeof(struct Packed));  /* 8 */
```

## Unions

```c
/* All members share the SAME memory */
typedef union {
    int   i;
    float f;
    char  bytes[4];
} Value;

Value v;
v.i = 0x3F800000;
printf("float: %f\\n", v.f);  /* 1.0 (IEEE 754 representation) */
printf("int:   %d\\n", v.i);  /* same bytes, different interpretation */
```

## Enumerations

```c
typedef enum {
    STATUS_PENDING  = 0,
    STATUS_ACTIVE   = 1,
    STATUS_CLOSED   = 2,
    STATUS_ARCHIVED = 3,
} Status;

Status s = STATUS_ACTIVE;
if (s == STATUS_ACTIVE) printf("Active\\n");
```

## Lab Exercise
1. Define a `struct Date` and write functions to compare and format dates
2. Measure struct padding: create optimally packed vs default struct, compare sizes
3. Implement a tagged union for a dynamic type system (int, float, string)
'''),

'_09_03_03_linked_list_implementation.md': ('09_03_03','Linked List Implementation','C',3,'Advanced C',3,'intermediate',['linked-list','node','head','tail','insert','delete','traverse','doubly-linked','malloc','free'],'''
## Singly Linked List

```c
#include <stdio.h>
#include <stdlib.h>

typedef struct Node {
    int data;
    struct Node *next;
} Node;

/* Create a new node */
Node *create_node(int data) {
    Node *node = malloc(sizeof(Node));
    if (!node) exit(EXIT_FAILURE);
    node->data = data;
    node->next = NULL;
    return node;
}

/* Insert at head */
Node *insert_head(Node *head, int data) {
    Node *node = create_node(data);
    node->next = head;
    return node;
}

/* Insert at tail */
Node *insert_tail(Node *head, int data) {
    Node *node = create_node(data);
    if (!head) return node;
    Node *curr = head;
    while (curr->next) curr = curr->next;
    curr->next = node;
    return head;
}

/* Delete by value */
Node *delete_node(Node *head, int data) {
    if (!head) return NULL;
    if (head->data == data) {
        Node *tmp = head->next;
        free(head);
        return tmp;
    }
    Node *curr = head;
    while (curr->next && curr->next->data != data)
        curr = curr->next;
    if (curr->next) {
        Node *tmp = curr->next->next;
        free(curr->next);
        curr->next = tmp;
    }
    return head;
}

/* Print */
void print_list(Node *head) {
    for (Node *curr = head; curr; curr = curr->next)
        printf("%d -> ", curr->data);
    printf("NULL\\n");
}

/* Free entire list */
void free_list(Node *head) {
    Node *curr = head;
    while (curr) {
        Node *tmp = curr->next;
        free(curr);
        curr = tmp;
    }
}

int main(void) {
    Node *list = NULL;
    list = insert_tail(list, 10);
    list = insert_tail(list, 20);
    list = insert_tail(list, 30);
    list = insert_head(list, 5);
    print_list(list);   /* 5 -> 10 -> 20 -> 30 -> NULL */
    list = delete_node(list, 20);
    print_list(list);   /* 5 -> 10 -> 30 -> NULL */
    free_list(list);
    return 0;
}
```

## Lab Exercise
1. Add `reverse_list()` that reverses in-place
2. Detect a cycle using Floyd's two-pointer algorithm
3. Implement a doubly linked list with `prev` pointer
'''),

'_09_04_01_file_io.md': ('09_04_01','File I/O','C',4,'Systems Programming',1,'intermediate',['fopen','fclose','fread','fwrite','fgets','fprintf','fseek','ftell','binary','text'],'''
## Text File I/O

```c
#include <stdio.h>

/* Write */
FILE *fp = fopen("data.txt", "w");   /* "w", "r", "a", "wb", "rb" */
if (!fp) { perror("fopen"); return 1; }

fprintf(fp, "Name: %s\\nAge: %d\\n", "Raja", 28);
fputs("Another line\\n", fp);
fclose(fp);

/* Read line by line */
fp = fopen("data.txt", "r");
char line[256];
while (fgets(line, sizeof(line), fp)) {
    line[strcspn(line, "\\n")] = '\\0';   /* remove newline */
    printf("[%s]\\n", line);
}
fclose(fp);
```

## Binary File I/O

```c
typedef struct { char name[50]; int age; float salary; } Employee;

Employee emp = {"Raja", 28, 75000.0f};

/* Write binary */
FILE *fp = fopen("data.bin", "wb");
fwrite(&emp, sizeof(Employee), 1, fp);
fclose(fp);

/* Read binary */
Employee loaded;
fp = fopen("data.bin", "rb");
fread(&loaded, sizeof(Employee), 1, fp);
fclose(fp);
printf("%s %d %.2f\\n", loaded.name, loaded.age, loaded.salary);
```

## File Position

```c
fp = fopen("data.bin", "rb");
fseek(fp, 0, SEEK_END);       /* go to end */
long size = ftell(fp);         /* get position = file size */
fseek(fp, 0, SEEK_SET);       /* back to start */
rewind(fp);                    /* also goes to start */

/* Random access — read 3rd record */
fseek(fp, 2 * sizeof(Employee), SEEK_SET);
fread(&loaded, sizeof(Employee), 1, fp);
```

## Error Handling

```c
if (fopen("missing.txt", "r") == NULL) {
    perror("fopen");   /* "fopen: No such file or directory" */
    /* or: fprintf(stderr, "Error: %s\\n", strerror(errno)); */
}
```

## Lab Exercise
1. Write a student grade book to a binary file, read it back, calculate class average
2. Implement `file_copy(src, dst)` that copies any file in 8KB chunks
3. Build a simple CSV reader that parses comma-separated values from a text file
'''),

'_09_04_02_preprocessor_and_macros.md': ('09_04_02','Preprocessor and Macros','C',4,'Systems Programming',2,'intermediate',['#define','#include','#ifdef','#ifndef','macro','function-macro','guard','#pragma','conditional-compilation'],'''
## Preprocessor Directives

```c
/* File inclusion */
#include <stdio.h>          /* system header */
#include "my_header.h"      /* local header */

/* Constants (macro) */
#define PI        3.14159265
#define MAX_SIZE  100
#define COMPANY   "TechCorp"

/* Function-like macros */
#define MAX(a, b) ((a) > (b) ? (a) : (b))
#define SQUARE(x) ((x) * (x))
#define ABS(x)    ((x) < 0 ? -(x) : (x))

/* Parentheses are essential! */
/* SQUARE(1+2) -> ((1+2)*(1+2)) = 9  (correct) */
/* without parens: (1+2*1+2) = 5  (wrong!) */
```

## Include Guards

```c
/* my_header.h */
#ifndef MY_HEADER_H
#define MY_HEADER_H

typedef struct { int x, y; } Point;
void point_print(Point p);

#endif /* MY_HEADER_H */

/* Or use #pragma once (non-standard but widely supported) */
#pragma once
```

## Conditional Compilation

```c
#define DEBUG 1

#if DEBUG
    #define LOG(msg) fprintf(stderr, "[DEBUG] %s:%d: %s\\n", __FILE__, __LINE__, msg)
#else
    #define LOG(msg) /* nothing */
#endif

/* Platform detection */
#ifdef _WIN32
    #define PATH_SEP "\\"
#else
    #define PATH_SEP "/"
#endif

/* gcc -DDEBUG=1 */
```

## Predefined Macros

```c
__FILE__    /* "main.c" */
__LINE__    /* 42 */
__func__    /* "main" */
__DATE__    /* "Jul 28 2024" */
__TIME__    /* "13:45:00" */
```

## Variadic Macros

```c
#define DEBUG_PRINT(fmt, ...) \\
    fprintf(stderr, "[%s:%d] " fmt "\\n", __FILE__, __LINE__, ##__VA_ARGS__)

DEBUG_PRINT("Value: %d", 42);
DEBUG_PRINT("Hello World");
```

## Lab Exercise
1. Write a `ASSERT(cond)` macro that prints file, line, and expression on failure
2. Create a debug logging system: DEBUG level off in release build via `-DNDEBUG`
3. Write a generic `SWAP(type, a, b)` macro and test for int, float, char
'''),

'_09_04_03_c_for_embedded_systems.md': ('09_04_03','C for Embedded Systems','C',4,'Systems Programming',3,'advanced',['embedded','microcontroller','volatile','register','bit-manipulation','memory-mapped-IO','interrupt','watchdog','MISRA'],'''
## Embedded C Concepts

```c
/* volatile — prevents compiler optimization for hardware registers */
volatile uint32_t *GPIO_PORT = (volatile uint32_t *)0x40020000;

/* Reading a hardware register (compiler won't cache this) */
uint32_t status = *GPIO_PORT;

/* Writing to hardware register */
*GPIO_PORT |= (1 << 5);   /* set bit 5 (enable pin) */
*GPIO_PORT &= ~(1 << 5);  /* clear bit 5 */
*GPIO_PORT ^= (1 << 5);   /* toggle bit 5 */
```

## Bit Manipulation

```c
#include <stdint.h>

uint8_t flags = 0;

/* Set bit n */
#define BIT_SET(reg, n)   ((reg) |=  (1U << (n)))
/* Clear bit n */
#define BIT_CLR(reg, n)   ((reg) &= ~(1U << (n)))
/* Toggle bit n */
#define BIT_TOG(reg, n)   ((reg) ^=  (1U << (n)))
/* Test bit n */
#define BIT_TST(reg, n)   (((reg) >> (n)) & 1U)

BIT_SET(flags, 3);    /* flags = 0b00001000 */
BIT_CLR(flags, 3);    /* flags = 0b00000000 */
```

## Fixed-Width Types (stdint.h)

```c
#include <stdint.h>

int8_t   a;   /* exactly 8-bit signed */
uint8_t  b;   /* exactly 8-bit unsigned */
int16_t  c;
uint16_t d;
int32_t  e;
uint32_t f;
int64_t  g;
uint64_t h;

/* Use these instead of int/long in embedded code */
```

## Memory-Mapped I/O Structure

```c
/* GPIO register map */
typedef struct {
    volatile uint32_t MODER;    /* offset 0x00 */
    volatile uint32_t OTYPER;   /* offset 0x04 */
    volatile uint32_t OSPEEDR;  /* offset 0x08 */
    volatile uint32_t PUPDR;    /* offset 0x0C */
    volatile uint32_t IDR;      /* offset 0x10 — input data */
    volatile uint32_t ODR;      /* offset 0x14 — output data */
    volatile uint32_t BSRR;     /* offset 0x18 — bit set/reset */
} GPIO_TypeDef;

#define GPIOA ((GPIO_TypeDef *)0x40020000)
GPIOA->ODR |= (1 << 5);   /* set pin 5 */
```

## Lab Exercise
1. Write a `ring_buffer_t` implementation for UART receive (used in ISR context)
2. Implement a software debounce for a button using a timer counter in C
3. Create a bit-field struct for a register map and verify its size equals the hardware spec
'''),

'_09_04_04_debugging_and_best_practices.md': ('09_04_04','Debugging and Best Practices','C',4,'Systems Programming',4,'intermediate',['gdb','valgrind','sanitizer','ASAN','assert','defensive-programming','MISRA','static-analysis','clang-tidy'],'''
## GDB Debugger

```bash
# Compile with debug info
gcc -g -O0 -o program program.c

# Start GDB
gdb ./program

# GDB commands
(gdb) run                # start program
(gdb) break main         # breakpoint at main
(gdb) break program.c:42 # breakpoint at line 42
(gdb) next               # next line (step over)
(gdb) step               # step into function
(gdb) continue           # continue to next breakpoint
(gdb) print x            # print variable x
(gdb) print *ptr         # dereference pointer
(gdb) info locals        # all local variables
(gdb) backtrace          # call stack
(gdb) quit
```

## AddressSanitizer (ASAN)

```bash
gcc -fsanitize=address -g -o program program.c
./program
# Reports: buffer overflow, heap use-after-free, stack overflow, leaks
```

## Static Analysis

```bash
# clang static analyzer
scan-build gcc -o program program.c

# cppcheck
cppcheck --enable=all program.c

# splint (MISRA-style)
splint program.c
```

## Defensive Programming

```c
#include <assert.h>

/* assert — checks during DEBUG, removed in release (-DNDEBUG) */
void array_set(int *arr, int idx, int val, int len) {
    assert(arr != NULL);
    assert(idx >= 0 && idx < len);
    arr[idx] = val;
}

/* Check all return values */
FILE *fp = fopen("file.txt", "r");
if (fp == NULL) {
    fprintf(stderr, "Cannot open file: %s\\n", strerror(errno));
    return -1;
}

/* Always null-check malloc */
int *buf = malloc(n * sizeof(int));
if (!buf) { perror("malloc"); exit(EXIT_FAILURE); }

/* Avoid magic numbers */
#define BUFFER_SIZE 4096
char buf[BUFFER_SIZE];
```

## Common C Bugs Checklist

| Bug | Prevention |
|---|---|
| Memory leak | Free every malloc, use valgrind |
| Buffer overflow | Use `snprintf`, `strncpy`, bounds check |
| Null dereference | Check before dereferencing |
| Uninitialized variable | Initialize all variables |
| Integer overflow | Use `UINT_MAX` checks |
| Off-by-one | Draw diagrams, test edge cases |
| Use after free | Set pointer to NULL after free |

## Lab Exercise
1. Find and fix 5 bugs in a provided buggy C program using GDB
2. Run the same program with ASAN — compare the errors caught
3. Write a `safe_malloc` wrapper that logs allocation size and always zero-inits
'''),
}

# Skip placeholder files
for fname, data in c_lessons.items():
    if data is None:
        # delete or overwrite placeholder
        path = os.path.join(BASE, C_DIR, fname)
        if os.path.exists(path):
            os.remove(path)
            print(f'  [DELETE] {fname}')
        continue
    lid, title, course, mod, mod_title, les, diff, tags, body = data
    write(C_DIR, fname, fm(lid, title, course, mod, mod_title, les, diff, tags) + body.strip() + '\n')

# ═══════════════════════════════════════════════════════════════
# C++ — 13 lessons
# ═══════════════════════════════════════════════════════════════
print()
print('='*60)
print('C++ — 13 lessons')
print('='*60)
CPP = '_10_cpp'

cpp_lessons = {
'_03_01_cpp_curriculum_placeholder.md': None,
'_10_01_01_cpp_overview_and_setup.md': ('10_01_01','C++ Overview and Setup','C++',1,'C++ Fundamentals',1,'beginner',['C++','g++','clang++','standard','C++11','C++17','C++20','CMake','RAII'],'''
## What is C++?

C++ is a **general-purpose, multi-paradigm** language (procedural, OOP, generic, functional) developed by Bjarne Stroustrup as a superset of C. Key features: **RAII**, zero-overhead abstractions, templates, STL.

## Standards Timeline

| Standard | Key Features |
|---|---|
| C++11 | Move semantics, lambdas, auto, range-for, `nullptr`, smart pointers |
| C++14 | Generic lambdas, return type deduction |
| C++17 | Structured bindings, `if constexpr`, `std::optional`, `std::variant` |
| C++20 | Concepts, ranges, coroutines, modules, `std::span` |
| C++23 | `std::expected`, `std::flat_map`, stackful coroutines |

## Setup

```bash
# Install GCC
sudo apt install g++ build-essential

# Compile
g++ -std=c++20 -Wall -Wextra -o hello hello.cpp
./hello

# CMakeLists.txt
cmake_minimum_required(VERSION 3.20)
project(MyApp CXX)
set(CMAKE_CXX_STANDARD 20)
add_executable(hello hello.cpp)
```

## Hello World

```cpp
#include <iostream>
#include <format>   // C++20

int main() {
    std::cout << "Hello, World!" << std::endl;
    std::cout << std::format("C++ is {} years old!\\n", 2024 - 1979);
    return 0;
}
```

## RAII Principle

**Resource Acquisition Is Initialization** — resources (memory, files, locks) are acquired in constructors and released in destructors automatically.

```cpp
// Classic C — must remember to close!
FILE *f = fopen("file.txt", "r");
// ... use f ...
fclose(f);

// C++ RAII — auto-closed on scope exit
{
    std::ifstream f("file.txt");
    // ... use f ...
}   // f.~ifstream() called — file closed!
```

## Lab Exercise
1. Write `hello.cpp` that uses `std::cout`, compile with `-std=c++20`
2. Create a `CMakeLists.txt` and build using `cmake .. && make`
3. Demonstrate RAII by creating a `TimedScope` class that measures execution time
'''),

'_10_01_02_references_and_value_types.md': ('10_01_02','References and Value Types','C++',1,'C++ Fundamentals',2,'intermediate',['reference','lvalue','rvalue','move-semantics','std::move','perfect-forwarding','auto','decltype'],'''
## References

```cpp
int x = 42;
int &ref = x;    // lvalue reference — must bind at declaration

ref = 100;       // modifies x!
std::cout << x;  // 100

// Const reference — does not allow modification
const int &cref = x;
// cref = 50;   // compile error

// Reference in function (avoids copy)
void double_value(int &n) { n *= 2; }
double_value(x);   // x is now 200

// Const reference parameter (safe, no copy)
void print(const std::string &s) { std::cout << s; }
```

## Value Categories: lvalue vs rvalue

```cpp
int x = 42;         // x is lvalue (has address)
42;                  // rvalue (temporary, no address)

// lvalue reference: int &r = x;   OK
// int &r = 42;    ERROR — can't bind lvalue-ref to rvalue

// rvalue reference (C++11)
int &&rref = 42;   // OK
int &&rref2 = std::move(x);  // move x as rvalue
```

## Move Semantics

```cpp
class Buffer {
    int *data;
    size_t size;
public:
    // Copy constructor — expensive!
    Buffer(const Buffer &other) : size(other.size) {
        data = new int[size];
        std::copy(other.data, other.data+size, data);
    }

    // Move constructor — cheap! (steal pointer)
    Buffer(Buffer &&other) noexcept
        : data(other.data), size(other.size) {
        other.data = nullptr;   // leave source empty
        other.size = 0;
    }

    ~Buffer() { delete[] data; }
};

Buffer a(1000);
Buffer b = std::move(a);   // move, not copy
// a.data is now nullptr — don't use a!
```

## auto and decltype

```cpp
auto x = 42;                // int
auto y = 3.14;              // double
auto z = std::string{"hi"}; // std::string

// auto in range-for
for (auto &item : container) { /* ... */ }
for (const auto &[key, value] : map) { /* C++17 structured binding */ }

// decltype — type of an expression without evaluating
decltype(x) copy = x;     // same type as x
auto result = compute();
decltype(result) backup;
```

## Lab Exercise
1. Show that passing `std::string` by value vs const-ref has different copy costs (with counter)
2. Implement a simple `UniqueResource<T>` with move constructor and deleted copy constructor
3. Use structured bindings to iterate a `std::map<std::string, int>`
'''),

'_10_01_03_functions_and_overloading.md': ('10_01_03','Functions and Overloading','C++',1,'C++ Fundamentals',3,'beginner',['overloading','default-arguments','inline','constexpr','lambda','std::function','trailing-return'],'''
## Function Overloading

```cpp
int    add(int a, int b)       { return a + b; }
double add(double a, double b) { return a + b; }
std::string add(const std::string &a, const std::string &b) { return a + b; }

add(1, 2)           // int version
add(1.5, 2.5)       // double version
add("Hello", " World")  // string version
```

## Default Arguments

```cpp
void connect(const std::string &host,
             int port = 8080,
             bool ssl = false) {
    std::cout << (ssl ? "https" : "http") << "://" << host << ":" << port;
}

connect("example.com");            // port=8080, ssl=false
connect("example.com", 443, true); // https://example.com:443
```

## constexpr Functions

```cpp
// Evaluated at compile-time when possible
constexpr int factorial(int n) {
    return n <= 1 ? 1 : n * factorial(n - 1);
}

constexpr int fact5 = factorial(5);  // computed at compile time
int arr[factorial(4)];               // OK: compile-time constant
```

## Lambda Expressions

```cpp
// [capture](params) -> return_type { body }
auto square = [](int x) { return x * x; };
square(5)   // 25

// Capture by value
int threshold = 10;
auto above = [threshold](int x) { return x > threshold; };

// Capture by reference
int count = 0;
auto counter = [&count]() { ++count; };

// Generic lambda (C++14)
auto add = [](auto a, auto b) { return a + b; };
add(1, 2)         // int
add(1.5, 2.5)     // double
add(std::string{"Hello"}, std::string{" World"})

// std::function
std::function<int(int, int)> op;
op = [](int a, int b) { return a + b; };
op = add_function;    // can hold any callable
```

## Lab Exercise
1. Overload `toString()` for int, double, bool, and vector
2. Write a `Timer` function using `std::function` that measures any callable
3. Implement `compose(f, g)` that returns a lambda `h(x) = f(g(x))`
'''),

'_10_02_01_classes_and_constructors.md': ('10_02_01','Classes and Constructors','C++',2,'Object-Oriented C++',1,'intermediate',['class','constructor','destructor','copy','move','initializer-list','explicit','default','delete'],'''
## Class Basics

```cpp
class BankAccount {
private:
    std::string owner_;
    double balance_;

public:
    // Constructor with initializer list (preferred)
    explicit BankAccount(std::string owner, double initial = 0.0)
        : owner_(std::move(owner)), balance_(initial) {
        if (initial < 0) throw std::invalid_argument("Negative balance");
    }

    // Copy constructor
    BankAccount(const BankAccount &other)
        : owner_(other.owner_), balance_(other.balance_) {}

    // Move constructor
    BankAccount(BankAccount &&other) noexcept
        : owner_(std::move(other.owner_)), balance_(other.balance_) {
        other.balance_ = 0;
    }

    // Destructor
    ~BankAccount() { /* clean up if needed */ }

    // Member functions
    void deposit(double amount) {
        if (amount <= 0) throw std::invalid_argument("Amount must be positive");
        balance_ += amount;
    }

    double balance() const { return balance_; }  // const: doesn't modify object

    // Friend function
    friend std::ostream &operator<<(std::ostream &os, const BankAccount &acc);
};

std::ostream &operator<<(std::ostream &os, const BankAccount &acc) {
    return os << "Account[" << acc.owner_ << ": $" << acc.balance_ << "]";
}
```

## Rule of Zero / Three / Five

```cpp
// Rule of Zero: if no custom destructor needed, use defaults
class Simple {
    std::string name;  // std::string manages its own memory
    int value = 0;
    // No custom destructor, copy, move needed!
};

// Rule of Five: if you define any of these, define all 5:
// destructor, copy constructor, copy assignment,
// move constructor, move assignment

// Delete to prevent copying
class NonCopyable {
public:
    NonCopyable() = default;
    NonCopyable(const NonCopyable &) = delete;
    NonCopyable &operator=(const NonCopyable &) = delete;
    NonCopyable(NonCopyable &&) = default;
    NonCopyable &operator=(NonCopyable &&) = default;
};
```

## Lab Exercise
1. Implement a `Matrix` class with constructor, copy/move, and `operator*`
2. Create a `RAII` file handle class that auto-closes on destruction
3. Demonstrate the difference between `explicit` and non-explicit constructors
'''),

'_10_02_02_operator_overloading.md': ('10_02_02','Operator Overloading','C++',2,'Object-Oriented C++',2,'intermediate',['operator+','operator<<','operator[]','operator==','operator<=>','spaceship','friend'],'''
## Overloading Arithmetic Operators

```cpp
struct Vector2D {
    double x, y;

    Vector2D(double x = 0, double y = 0) : x(x), y(y) {}

    // Member operator
    Vector2D operator+(const Vector2D &rhs) const {
        return {x + rhs.x, y + rhs.y};
    }

    Vector2D &operator+=(const Vector2D &rhs) {
        x += rhs.x; y += rhs.y; return *this;
    }

    // Scalar multiplication
    Vector2D operator*(double scalar) const { return {x*scalar, y*scalar}; }

    // Unary minus
    Vector2D operator-() const { return {-x, -y}; }

    // Length
    double length() const { return std::sqrt(x*x + y*y); }

    // Comparison (C++20 spaceship operator)
    auto operator<=>(const Vector2D &) const = default;  // lexicographic
    bool operator==(const Vector2D &) const = default;
};

// Non-member: double * Vector2D
Vector2D operator*(double scalar, const Vector2D &v) { return v * scalar; }

// Stream output
std::ostream &operator<<(std::ostream &os, const Vector2D &v) {
    return os << "(" << v.x << ", " << v.y << ")";
}

Vector2D a{1, 2}, b{3, 4};
auto c = a + b;           // (4, 6)
std::cout << c;
auto d = 2.0 * a;         // (2, 4)
bool eq = (a == a);       // true
```

## Subscript and Call Operators

```cpp
class Matrix {
    std::vector<std::vector<double>> data_;
    int rows_, cols_;
public:
    // operator[] for row access
    std::vector<double> &operator[](int i) { return data_[i]; }
    const std::vector<double> &operator[](int i) const { return data_[i]; }

    // operator() for element access
    double &operator()(int r, int c) { return data_[r][c]; }
};

Matrix m(3, 3);
m[0][0] = 1.0;
m(1, 2) = 3.14;
```

## Lab Exercise
1. Implement a `Fraction` class with `+`, `-`, `*`, `/`, `<<`, `==` operators
2. Add `<=>` spaceship operator to `Fraction` and verify sorting works
3. Create a `JSON` class with `operator[]` for string keys and integer indices
'''),

'_10_02_03_inheritance_and_polymorphism.md': ('10_02_03','Inheritance and Polymorphism','C++',2,'Object-Oriented C++',3,'intermediate',['virtual','override','final','pure-virtual','abstract','vtable','dynamic_cast','RTTI'],'''
## Virtual Functions and Polymorphism

```cpp
class Shape {
public:
    virtual double area() const = 0;       // pure virtual
    virtual double perimeter() const = 0;  // pure virtual
    virtual std::string name() const { return "Shape"; }
    virtual ~Shape() = default;            // virtual destructor!
};

class Circle : public Shape {
    double radius_;
public:
    explicit Circle(double r) : radius_(r) {}
    double area()      const override { return M_PI * radius_ * radius_; }
    double perimeter() const override { return 2 * M_PI * radius_; }
    std::string name() const override { return "Circle"; }
};

// Polymorphism via pointer/reference
std::vector<std::unique_ptr<Shape>> shapes;
shapes.push_back(std::make_unique<Circle>(5));
shapes.push_back(std::make_unique<Rectangle>(4, 6));

for (const auto &s : shapes) {
    std::cout << s->name() << ": area=" << s->area() << "\\n";
}
```

## dynamic_cast and RTTI

```cpp
Shape *ptr = get_some_shape();

// Safe downcast — returns nullptr on failure
if (auto *c = dynamic_cast<Circle *>(ptr)) {
    std::cout << "It's a circle with r=" << c->radius() << "\\n";
}

// typeid
std::cout << typeid(*ptr).name() << "\\n";
```

## final and override

```cpp
class Animal {
    virtual void speak() const {}
};

class Dog final : public Animal {   // cannot be subclassed
    void speak() const override final { std::cout << "Woof!\\n"; }
};
```

## Lab Exercise
1. Build a `Shape` hierarchy, compute total area polymorphically
2. Demonstrate why `virtual ~Shape()` is essential — show double-free bug without it
3. Use `dynamic_cast` to safely downcast a base pointer to derived
'''),

'_10_03_01_smart_pointers.md': ('10_03_01','Smart Pointers','C++',3,'Modern C++ Memory',1,'intermediate',['unique_ptr','shared_ptr','weak_ptr','make_unique','make_shared','RAII','ownership','custom-deleter'],'''
## unique_ptr — Exclusive Ownership

```cpp
#include <memory>

// Create
auto p = std::make_unique<int>(42);         // preferred
auto arr = std::make_unique<int[]>(100);    // array

// Use
std::cout << *p << "\\n";   // 42
*p = 100;

// Transfer ownership (move only — not copyable)
auto p2 = std::move(p);    // p is now nullptr
// p  = nullptr
// p2 = owns the int

// Custom deleter
auto file = std::unique_ptr<FILE, decltype(&fclose)>(
    fopen("file.txt", "r"), fclose
);
```

## shared_ptr — Shared Ownership

```cpp
auto sp1 = std::make_shared<std::string>("shared data");
auto sp2 = sp1;     // ref count = 2
auto sp3 = sp1;     // ref count = 3

sp1.use_count()     // 3
sp1.reset();        // ref count = 2
// sp2 and sp3 still valid
// Memory freed when last shared_ptr destroyed

// shared_ptr in container
std::vector<std::shared_ptr<Node>> nodes;
nodes.push_back(std::make_shared<Node>(1));
```

## weak_ptr — Non-Owning Reference

```cpp
// Breaks cycles! (parent-child cycle would leak with shared_ptr)
struct Node {
    int value;
    std::shared_ptr<Node> next;
    std::weak_ptr<Node> prev;   // non-owning back-pointer
};

// Use weak_ptr
std::weak_ptr<Widget> weak = shared_ptr_widget;
if (auto locked = weak.lock()) {  // lock() returns shared_ptr or nullptr
    locked->doSomething();
} else {
    std::cout << "Object destroyed!\\n";
}
```

## Ownership Guidelines

| Need | Tool |
|---|---|
| Single owner | `unique_ptr` |
| Shared ownership | `shared_ptr` |
| Non-owning reference to shared | `weak_ptr` |
| Stack object | Direct (no pointer) |
| Raw pointer | Observer/non-owning only |
| Raw `new` | Avoid! |

## Lab Exercise
1. Implement a linked list using `unique_ptr<Node>` for ownership
2. Show a `shared_ptr` cycle leak, fix it with `weak_ptr`
3. Build a tree structure with `shared_ptr` children and `weak_ptr` parent
'''),

'_10_03_02_templates.md': ('10_03_02','Templates','C++',3,'Modern C++ Memory',2,'advanced',['template','class-template','function-template','specialization','SFINAE','concepts','variadic-template','type-traits'],'''
## Function Templates

```cpp
template <typename T>
T max_val(T a, T b) {
    return (a > b) ? a : b;
}

max_val(3, 7)            // int
max_val(3.14, 2.71)      // double
max_val<float>(1.5f, 2.5f)  // explicit instantiation

// Multiple type parameters
template <typename T, typename U>
auto add(T a, U b) -> decltype(a + b) {
    return a + b;
}
```

## Class Templates

```cpp
template <typename T, int Capacity = 10>
class FixedStack {
    T data_[Capacity];
    int top_ = 0;
public:
    void push(const T &val) {
        if (top_ >= Capacity) throw std::overflow_error("Stack full");
        data_[top_++] = val;
    }

    T pop() {
        if (top_ == 0) throw std::underflow_error("Stack empty");
        return data_[--top_];
    }

    int size() const { return top_; }
    bool empty() const { return top_ == 0; }
};

FixedStack<int, 5> istack;
FixedStack<std::string> sstack;   // uses default capacity 10
```

## Concepts (C++20)

```cpp
// Define concept
template <typename T>
concept Numeric = std::is_arithmetic_v<T>;

template <typename T>
concept Sortable = requires(T &container) {
    container.begin();
    container.end();
    { *container.begin() } -> std::totally_ordered;
};

// Use in function
template <Numeric T>
T square(T x) { return x * x; }

square(5)      // OK
square("hi")   // compile error — "string" does not satisfy Numeric
```

## Variadic Templates

```cpp
// Base case
void print() {}

// Recursive case
template <typename T, typename... Rest>
void print(T first, Rest... rest) {
    std::cout << first;
    if constexpr (sizeof...(rest) > 0) std::cout << ", ";
    print(rest...);
}

print(1, 3.14, "hello", true);   // 1, 3.14, hello, 1
```

## Lab Exercise
1. Implement a generic `Pair<T,U>` with `swap()`, `operator==`, and `make_pair`
2. Write a `TypeList<Types...>` that computes `size` and `get<N>` at compile time
3. Use Concepts to constrain a `BinaryTree<T>` to only accept `std::totally_ordered` types
'''),

'_10_03_03_stl_containers_and_algorithms.md': ('10_03_03','STL Containers and Algorithms','C++',3,'Modern C++ Memory',3,'intermediate',['vector','map','set','unordered_map','array','deque','list','algorithm','sort','find','transform','ranges'],'''
## Key STL Containers

```cpp
#include <vector>
#include <map>
#include <set>
#include <unordered_map>
#include <array>
#include <deque>

std::vector<int> v = {3, 1, 4, 1, 5, 9};
v.push_back(2);
v.emplace_back(6);             // construct in-place
v.reserve(100);                // pre-allocate
v.size(); v.empty(); v.front(); v.back();

std::map<std::string, int> m;   // sorted by key O(log n)
m["Alice"] = 95;
m.find("Alice")               // iterator or end()
m.count("Bob")                // 0 or 1
m.emplace("Carol", 88);

std::unordered_map<std::string, int> um;  // O(1) average
um.reserve(100);

std::set<int> s = {3,1,4,1,5};  // sorted unique: {1,3,4,5}
s.insert(9);
s.count(1);   // 1 or 0

std::array<int, 5> arr = {1,2,3,4,5};   // fixed-size, no heap
```

## Algorithms

```cpp
#include <algorithm>
#include <numeric>

std::vector<int> v = {3,1,4,1,5,9,2,6};

std::sort(v.begin(), v.end());                 // ascending
std::sort(v.begin(), v.end(), std::greater{}); // descending

std::find(v.begin(), v.end(), 5);  // iterator to 5
std::count(v.begin(), v.end(), 1); // 2
std::binary_search(v.begin(), v.end(), 9); // true (must be sorted)

std::transform(v.begin(), v.end(), v.begin(), [](int x){ return x*x; });

std::accumulate(v.begin(), v.end(), 0);        // sum
std::reduce(v.begin(), v.end());               // faster, parallel-friendly

std::partition(v.begin(), v.end(), [](int x){ return x%2==0; });
std::remove_if(v.begin(), v.end(), [](int x){ return x<3; }); // + erase!
v.erase(std::remove_if(v.begin(), v.end(), pred), v.end());    // erase-remove
```

## Ranges (C++20)

```cpp
#include <ranges>
namespace rv = std::views;

std::vector<int> nums = {1,2,3,4,5,6,7,8,9,10};

// Lazy pipeline
auto result = nums
    | rv::filter([](int n){ return n % 2 == 0; })
    | rv::transform([](int n){ return n * n; })
    | rv::take(3);

for (int n : result) std::cout << n << " ";  // 4 16 36
```

## Lab Exercise
1. Find the top 5 most frequent words in a text using `unordered_map` + `partial_sort`
2. Implement a priority queue using `std::vector` + `std::make_heap`
3. Rewrite a series of loops as a ranges pipeline
'''),

'_10_04_01_cpp11_to_cpp23_features.md': ('10_04_01','C++11 to C++23 Features','C++',4,'Modern C++',1,'intermediate',['auto','constexpr','nullptr','range-for','initializer_list','optional','variant','any','expected','span'],'''
## C++11 Essentials

```cpp
// auto type deduction
auto x = 42;
auto it = container.begin();

// Range-based for
for (const auto &item : container) { /* ... */ }

// nullptr (replaces NULL)
int *p = nullptr;

// Initializer lists
std::vector<int> v = {1, 2, 3, 4, 5};

// Move semantics — std::move, std::forward
// Lambda expressions — [capture](params){ body }
// constexpr — compile-time evaluation
// Smart pointers — unique_ptr, shared_ptr, weak_ptr
```

## C++17 Features

```cpp
// std::optional — nullable value without pointer
std::optional<int> find_value(int key) {
    if (auto it = map.find(key); it != map.end())
        return it->second;
    return std::nullopt;
}

auto val = find_value(42);
if (val) std::cout << *val;
val.value_or(0);   // default if nullopt

// std::variant — type-safe union
std::variant<int, double, std::string> v;
v = 42;
v = 3.14;
v = "hello";
std::get<std::string>(v)       // "hello"
std::holds_alternative<int>(v) // false
std::visit([](auto &&val){ std::cout << val; }, v);

// Structured bindings
auto [key, value] = *map.begin();
auto [x, y, z] = std::tuple{1, 2.0, "three"};

// if/switch with initializer
if (auto it = m.find("key"); it != m.end()) {
    use(it->second);
}
```

## C++20 Features

```cpp
// Concepts
template <std::integral T>
T gcd(T a, T b) { return b ? gcd(b, a % b) : a; }

// std::span — non-owning view of contiguous data
void process(std::span<const int> data) {
    for (int x : data) { /* ... */ }
}
process(std::vector<int>{1,2,3});
process(std::array<int,3>{1,2,3});

// Ranges
auto evens = std::views::iota(1, 101)
           | std::views::filter([](int n){ return n%2==0; });
```

## C++23 Features

```cpp
// std::expected — error handling without exceptions
std::expected<int, std::string> parse_int(std::string_view s) {
    try { return std::stoi(std::string{s}); }
    catch (...) { return std::unexpected("Not a number: " + std::string{s}); }
}

auto result = parse_int("42");
if (result) std::cout << *result;    // 42
else std::cout << result.error();   // "Not a number: ..."
```

## Lab Exercise
1. Replace raw pointer with `optional<T>` in a find function, handle nullopt gracefully
2. Implement a `Result<T, E>` type using `std::variant` (before C++23)
3. Process a data pipeline using C++20 ranges: filter → transform → take → to vector
'''),

'_10_04_02_concurrency_in_cpp.md': ('10_04_02','Concurrency in C++','C++',4,'Modern C++',2,'advanced',['thread','mutex','lock_guard','atomic','future','promise','async','condition_variable','thread_pool'],'''
## std::thread

```cpp
#include <thread>
#include <mutex>
#include <atomic>
#include <future>

void task(int id) {
    std::cout << "Thread " << id << "\\n";
}

std::thread t1(task, 1);
std::thread t2(task, 2);
t1.join();   // wait for t1
t2.join();
```

## Mutex and Lock

```cpp
std::mutex mtx;
int shared = 0;

void increment() {
    std::lock_guard<std::mutex> lock(mtx);  // RAII — auto unlock
    shared++;
}

// unique_lock — flexible (can unlock manually)
std::unique_lock<std::mutex> lock(mtx);
lock.unlock();
// do other work...
lock.lock();
```

## Atomic Operations

```cpp
std::atomic<int> counter = 0;
counter.fetch_add(1, std::memory_order_relaxed);
counter++;    // operator++ is atomic
int val = counter.load();
counter.store(0);
counter.compare_exchange_strong(expected, desired);
```

## std::async and std::future

```cpp
// Launch async task (may run in new thread)
std::future<int> fut = std::async(std::launch::async, []() {
    return expensive_computation();
});

// Do other work...
int result = fut.get();   // blocks until done
```

## condition_variable

```cpp
std::mutex mtx;
std::condition_variable cv;
bool ready = false;

// Producer
{
    std::lock_guard<std::mutex> lock(mtx);
    ready = true;
}
cv.notify_one();

// Consumer
{
    std::unique_lock<std::mutex> lock(mtx);
    cv.wait(lock, [&]{ return ready; });
    // proceed
}
```

## Lab Exercise
1. Build a thread-safe queue using `mutex` + `condition_variable`
2. Parallelize a Monte Carlo π estimation with `std::async` across N threads
3. Implement a simple thread pool that executes `std::function<void()>` tasks
'''),

'_10_04_03_cpp_for_embedded_and_performance.md': ('10_04_03','C++ for Embedded and Performance','C++',4,'Modern C++',3,'advanced',['embedded','no-RTTI','no-exceptions','placement-new','constexpr','SIMD','profiling','benchmark'],'''
## Embedded C++ Constraints

Embedded systems often restrict:
- **Exceptions** (use error codes or `std::expected`)
- **RTTI** / `dynamic_cast` (use tag-based dispatch)
- **Dynamic allocation** (use static/stack allocation)
- **Standard library** (MCU has no OS, limited RAM)

## Compile Options for Embedded

```cmake
target_compile_options(firmware PRIVATE
    -fno-exceptions
    -fno-rtti
    -fno-unwind-tables
    -Os                    # optimize for size
    -march=armv7-m
    -mthumb
)
```

## Stack/Static Allocation

```cpp
// Static buffer instead of std::vector
template <typename T, std::size_t N>
class StaticVector {
    std::array<T, N> data_;
    std::size_t size_ = 0;
public:
    void push_back(const T &val) {
        if (size_ >= N) return;  // silently drop or assert
        data_[size_++] = val;
    }
    T &operator[](std::size_t i) { return data_[i]; }
    std::size_t size() const { return size_; }
};

StaticVector<int, 64> buf;   // no heap!
```

## constexpr for Compile-Time Tables

```cpp
// Lookup table computed at compile time
constexpr std::array<uint8_t, 256> make_crc_table() {
    std::array<uint8_t, 256> table{};
    for (int i = 0; i < 256; i++) {
        uint8_t crc = i;
        for (int j = 0; j < 8; j++)
            crc = (crc & 1) ? (crc >> 1) ^ 0x8C : (crc >> 1);
        table[i] = crc;
    }
    return table;
}
constexpr auto CRC_TABLE = make_crc_table();
```

## Performance Profiling

```bash
# GCC with gprof
g++ -pg -O2 -o program program.cpp
./program
gprof program gmon.out | head -30

# Valgrind cachegrind
valgrind --tool=cachegrind ./program

# Perf (Linux)
perf stat ./program
perf record ./program && perf report
```

## Lab Exercise
1. Implement `StaticVector<T, N>` and `StaticQueue<T, N>` for embedded use
2. Generate a sine LUT using `constexpr` (avoid runtime floating-point on MCU)
3. Profile a matrix multiplication: compare -O0 vs -O3 vs -O3 + SIMD
'''),
}

for fname, data in cpp_lessons.items():
    if data is None:
        path = os.path.join(BASE, CPP, fname)
        if os.path.exists(path):
            os.remove(path)
            print(f'  [DELETE] {fname}')
        continue
    lid, title, course, mod, mod_title, les, diff, tags, body = data
    write(CPP, fname, fm(lid, title, course, mod, mod_title, les, diff, tags) + body.strip() + '\n')

print()
print('='*60)
print(f'PHASE 4 COMPLETE — Total files written: {written}')
print('='*60)
